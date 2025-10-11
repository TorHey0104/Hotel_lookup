"""Tkinter GUI to manage Excel to JSON helper configuration."""

from __future__ import annotations

import json
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from pathlib import Path
from typing import Dict, List

from .excel_helper_config import (
    ExcelHelperConfigEntry,
    ExcelHelperConfigStore,
    detect_email_headers,
)


class ExcelHelperWindow:
    """Modal dialog that allows configuring Excel column selections."""

    def __init__(self, parent: tk.Tk, config_path: Path):
        self.parent = parent
        self.config_path = config_path
        self.config_store = ExcelHelperConfigStore(config_path)

        self.window = tk.Toplevel(parent)
        self.window.title("Excel Helper")
        self.window.geometry("780x560")
        self.window.transient(parent)
        self.window.grab_set()

        self.selected_excel: Path | None = None
        self.column_vars: Dict[str, tk.BooleanVar] = {}
        self.email_vars: Dict[str, tk.BooleanVar] = {}
        self.saved_var = tk.StringVar()

        self._build_ui()
        self._refresh_saved_entries()
        self._load_last_used_entry()

    def _build_ui(self) -> None:
        container = ttk.Frame(self.window, padding=16)
        container.grid(row=0, column=0, sticky="nsew")
        self.window.columnconfigure(0, weight=1)
        self.window.rowconfigure(0, weight=1)

        ttk.Label(
            container,
            text="Wählen Sie eine Excel-Datei aus und markieren Sie die relevanten Spalten.",
            wraplength=640,
        ).grid(row=0, column=0, columnspan=3, sticky="w")

        select_button = ttk.Button(container, text="Excel-Datei auswählen", command=self._choose_excel_file)
        select_button.grid(row=1, column=0, sticky="w", pady=(12, 12))

        self.file_label = ttk.Label(container, text="Keine Datei gewählt", foreground="#555555")
        self.file_label.grid(row=1, column=1, columnspan=2, sticky="w", padx=(12, 0))

        ttk.Label(container, text="Gespeicherte Konfiguration laden:").grid(
            row=2, column=0, sticky="w", pady=(0, 4)
        )
        self.saved_combo = ttk.Combobox(
            container,
            textvariable=self.saved_var,
            state="disabled",
            width=80,
        )
        self.saved_combo.grid(row=3, column=0, columnspan=3, sticky="we", pady=(0, 12))
        self.saved_combo.bind("<<ComboboxSelected>>", lambda _event: self._on_saved_selection())

        self.columns_frame = ttk.LabelFrame(container, text="Spaltenauswahl", padding=12)
        self.columns_frame.grid(row=4, column=0, columnspan=3, sticky="nsew")
        self.columns_inner = self._create_scrollable_section(self.columns_frame)

        self.contacts_frame = ttk.LabelFrame(
            container, text="E-Mail-Spalten für Drafts", padding=12
        )
        self.contacts_frame.grid(row=5, column=0, columnspan=3, sticky="nsew", pady=(12, 0))
        self.contacts_inner = self._create_scrollable_section(self.contacts_frame)

        button_frame = ttk.Frame(container)
        button_frame.grid(row=6, column=0, columnspan=3, sticky="ew", pady=(12, 0))

        self.save_button = ttk.Button(button_frame, text="Konfiguration speichern", command=self._save_config)
        self.save_button.grid(row=0, column=0, sticky="w")
        self.save_button.configure(state=tk.DISABLED)

        preview_button = ttk.Button(button_frame, text="Konfiguration anzeigen", command=self._show_preview)
        preview_button.grid(row=0, column=1, sticky="w", padx=(12, 0))

        close_button = ttk.Button(button_frame, text="Schließen", command=self.window.destroy)
        close_button.grid(row=0, column=2, sticky="e")

        self.preview_text = tk.Text(container, height=8, width=80)
        self.preview_text.grid(row=7, column=0, columnspan=3, sticky="nsew", pady=(12, 0))
        self.preview_text.configure(state="disabled")

        info_label = ttk.Label(
            container,
            text=f"Konfigurationsdatei: {self.config_path}",
            foreground="#555555",
        )
        info_label.grid(row=8, column=0, columnspan=3, sticky="w", pady=(8, 0))

        container.rowconfigure(4, weight=1)
        container.rowconfigure(5, weight=1)
        container.rowconfigure(7, weight=1)

        self.window.bind("<Escape>", lambda _event: self.window.destroy())

    def _create_scrollable_section(self, frame: ttk.LabelFrame) -> ttk.Frame:
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(0, weight=1)

        canvas = tk.Canvas(frame, borderwidth=0, highlightthickness=0)
        canvas.grid(row=0, column=0, sticky="nsew")

        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        canvas.configure(yscrollcommand=scrollbar.set)

        inner = ttk.Frame(canvas)
        canvas.create_window((0, 0), window=inner, anchor="nw")

        def _on_configure(event: tk.Event) -> None:
            canvas.configure(scrollregion=canvas.bbox("all"))

        inner.bind("<Configure>", _on_configure)
        canvas.bind(
            "<Enter>",
            lambda _event: canvas.bind_all("<MouseWheel>", lambda e: self._on_mousewheel(e, canvas)),
        )
        canvas.bind(
            "<Leave>",
            lambda _event: canvas.unbind_all("<MouseWheel>"),
        )
        canvas.bind("<Button-4>", lambda e: self._on_mousewheel(e, canvas))
        canvas.bind("<Button-5>", lambda e: self._on_mousewheel(e, canvas))

        return inner

    @staticmethod
    def _on_mousewheel(event: tk.Event, canvas: tk.Canvas) -> None:
        if getattr(event, "delta", 0):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        elif getattr(event, "num", None) == 4:
            canvas.yview_scroll(-1, "units")
        elif getattr(event, "num", None) == 5:
            canvas.yview_scroll(1, "units")

    def _clear_frames(self) -> None:
        for container in (self.columns_inner, self.contacts_inner):
            for child in container.winfo_children():
                child.destroy()
        self.column_vars.clear()
        self.email_vars.clear()

    def _choose_excel_file(self) -> None:
        file_path = filedialog.askopenfilename(
            parent=self.window,
            title="Excel-Datei auswählen",
            filetypes=[("Excel-Dateien", "*.xlsx"), ("Alle Dateien", "*.*")],
        )
        if not file_path:
            return
        excel_path = Path(file_path)
        if not excel_path.exists():
            messagebox.showerror("Datei nicht gefunden", f"Die Datei '{excel_path}' konnte nicht geöffnet werden.")
            return
        try:
            headers = self._load_headers(excel_path)
        except ModuleNotFoundError:
            messagebox.showerror(
                "openpyxl nicht installiert",
                "Für den Excel-Import wird 'openpyxl' benötigt. Installiere das Paket z. B. mit `pip install openpyxl`.",
            )
            return
        except ValueError as exc:
            messagebox.showerror("Fehler beim Lesen", str(exc))
            return

        self.selected_excel = excel_path
        self.file_label.configure(text=str(excel_path))
        self._populate_headers(headers)
        self.save_button.configure(state=tk.NORMAL)
        self._show_preview()
        self.saved_var.set(str(excel_path))
        self._refresh_saved_entries()

    def _load_headers(self, excel_path: Path) -> List[str]:
        try:
            from openpyxl import load_workbook  # type: ignore
        except ModuleNotFoundError as exc:
            raise exc
        workbook = load_workbook(excel_path, data_only=True, read_only=True)
        worksheet = workbook.active
        first_row = next(worksheet.iter_rows(values_only=True), None)
        if not first_row:
            raise ValueError("Die Arbeitsmappe enthält keine Daten.")
        headers = [str(cell).strip() if cell is not None else "" for cell in first_row]
        return headers

    def _populate_headers(self, headers: List[str]) -> None:
        self._clear_frames()
        if not headers:
            ttk.Label(self.columns_frame, text="Keine Spaltenüberschriften gefunden.").grid(sticky="w")
            ttk.Label(self.contacts_frame, text="Keine E-Mail-Spalten gefunden.").grid(sticky="w")
            return

        stored = self.config_store.get_entry(self.selected_excel) if self.selected_excel else None
        stored_columns = set(stored.selected_columns) if stored else None
        stored_emails = set(stored.email_columns) if stored else None
        detected_emails = set(detect_email_headers(headers))

        for idx, header in enumerate(headers):
            if not header:
                continue
            default_selected = True if stored_columns is None else header in stored_columns
            var = tk.BooleanVar(value=default_selected)
            check = ttk.Checkbutton(self.columns_inner, text=header, variable=var)
            check.grid(row=idx, column=0, sticky="w", pady=2)
            self.column_vars[header] = var
            var.trace_add("write", lambda *_args, header=header: self._on_column_toggle(header))

            email_default = False
            if stored_emails is None:
                email_default = header in detected_emails
            else:
                email_default = header in stored_emails
            self.email_vars[header] = tk.BooleanVar(value=email_default)

        self._refresh_email_options()

    def _on_column_toggle(self, header: str) -> None:
        column_var = self.column_vars.get(header)
        email_var = self.email_vars.get(header)
        if column_var is None or email_var is None:
            return
        if not column_var.get():
            email_var.set(False)
        self._refresh_email_options()

    def _refresh_email_options(self) -> None:
        for child in self.contacts_inner.winfo_children():
            child.destroy()

        selected_headers = [header for header, var in self.column_vars.items() if var.get()]
        if not selected_headers:
            ttk.Label(self.contacts_inner, text="Bitte wählen Sie zunächst Spalten aus.").grid(sticky="w")
            return

        for idx, header in enumerate(selected_headers):
            email_var = self.email_vars.setdefault(header, tk.BooleanVar(value=False))
            ttk.Checkbutton(self.contacts_inner, text=header, variable=email_var).grid(
                row=idx,
                column=0,
                sticky="w",
                pady=2,
            )

    def _save_config(self) -> None:
        if not self.selected_excel:
            messagebox.showinfo("Keine Datei", "Bitte wählen Sie zuerst eine Excel-Datei aus.")
            return
        selected_columns = [header for header, var in self.column_vars.items() if var.get()]
        email_columns = []
        for header, email_var in self.email_vars.items():
            column_var = self.column_vars.get(header)
            if column_var and column_var.get() and email_var.get():
                email_columns.append(header)
        if not selected_columns:
            messagebox.showwarning("Keine Spalten gewählt", "Bitte wählen Sie mindestens eine Spalte aus.")
            return
        entry = self.config_store.save_entry(self.selected_excel, selected_columns, email_columns)
        messagebox.showinfo(
            "Gespeichert",
            "Die Konfiguration wurde gespeichert und kann künftig über den Excel Helper erneut geladen werden.\n"
            f"Ablage: {self.config_path}\n"
            "Nutzen Sie die Auswahl anschließend mit dem Skript 'tools/excel_to_fixture.py', um eine Fixture-Datei zu erzeugen.",
        )
        self._update_preview(entry)
        self._refresh_saved_entries()

    def _show_preview(self) -> None:
        if not self.selected_excel:
            self._set_preview_text("{}")
            return
        pretty = self.config_store.to_pretty_json(self.selected_excel)
        self._set_preview_text(pretty)

    def _update_preview(self, entry: ExcelHelperConfigEntry) -> None:
        if not self.selected_excel:
            return
        payload = {
            "excelPath": str(self.selected_excel),
            **entry.to_dict(),
        }
        self._set_preview_text(json.dumps(payload, indent=2, ensure_ascii=False))

    def _refresh_saved_entries(self) -> None:
        entries = [str(path) for path in self.config_store.list_entries()]
        if entries:
            self.saved_combo.configure(state="readonly")
            self.saved_combo["values"] = entries
        else:
            self.saved_combo.configure(state="disabled")
            self.saved_combo["values"] = []
            self.saved_var.set("")

    def _on_saved_selection(self) -> None:
        selection = self.saved_var.get().strip()
        if not selection:
            return
        excel_path = Path(selection)
        if excel_path.exists():
            self._load_entry_from_path(excel_path)
        else:
            messagebox.showwarning(
                "Datei nicht gefunden",
                "Die gespeicherte Excel-Datei wurde nicht gefunden. Die Vorschau zeigt die zuletzt gespeicherte Konfiguration.",
            )
            entry = self.config_store.get_entry(excel_path)
            self.selected_excel = excel_path
            self.file_label.configure(text=f"{excel_path} (nicht gefunden)")
            self._clear_frames()
            if entry:
                payload = {
                    "excelPath": str(excel_path),
                    **entry.to_dict(),
                }
                self._set_preview_text(json.dumps(payload, indent=2, ensure_ascii=False))
            else:
                self._set_preview_text("{}")
            self.save_button.configure(state=tk.DISABLED)

    def _load_entry_from_path(self, excel_path: Path, *, silent: bool = False) -> None:
        try:
            headers = self._load_headers(excel_path)
        except ModuleNotFoundError:
            if not silent:
                messagebox.showerror(
                    "openpyxl nicht installiert",
                    "Für den Excel-Import wird 'openpyxl' benötigt. Installiere das Paket z. B. mit `pip install openpyxl`.",
                )
            return
        except ValueError as exc:
            if not silent:
                messagebox.showerror("Fehler beim Lesen", str(exc))
            return

        self.selected_excel = excel_path
        self.file_label.configure(text=str(excel_path))
        self._populate_headers(headers)
        self.save_button.configure(state=tk.NORMAL)
        self._show_preview()

    def _load_last_used_entry(self) -> None:
        last_used = self.config_store.get_last_used_path()
        if not last_used:
            return
        self.saved_var.set(str(last_used))
        if last_used.exists():
            self._load_entry_from_path(last_used, silent=True)
        else:
            entry = self.config_store.get_entry(last_used)
            if entry:
                payload = {
                    "excelPath": str(last_used),
                    **entry.to_dict(),
                }
                self._set_preview_text(json.dumps(payload, indent=2, ensure_ascii=False))
                self.file_label.configure(text=f"{last_used} (nicht gefunden)")
                self.save_button.configure(state=tk.DISABLED)

    def _set_preview_text(self, content: str) -> None:
        self.preview_text.configure(state="normal")
        self.preview_text.delete("1.0", tk.END)
        self.preview_text.insert("1.0", content)
        self.preview_text.configure(state="disabled")


def open_excel_helper(parent: tk.Tk, config_path: Path) -> ExcelHelperWindow:
    """Open the Excel helper dialog and wait until it is closed."""

    window = ExcelHelperWindow(parent, config_path)
    parent.wait_window(window.window)
    return window
