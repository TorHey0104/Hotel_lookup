"""Tkinter UI for the Spirit Lookup application."""

from __future__ import annotations

import json
import re
import tkinter as tk
from dataclasses import dataclass
from pathlib import Path
from types import ModuleType
from tkinter import filedialog, messagebox, ttk
from typing import List

from .config import AppConfig
from .controller import LookupResult, SpiritLookupController
from .display_config import DisplayConfig, DisplayFieldDefinition
from .excel_helper import open_excel_helper
from .excel_helper_config import ExcelHelperConfigStore, normalize_header
from .mail import MailClientError, open_mail_client
from .models import SpiritRecord
from .providers import DataProviderError, RecordNotFoundError


@dataclass
class RenderedField:
    """Represents a value that should be rendered in the detail dialog."""

    label: str
    value: str | None
    is_email: bool


_CONTACT_LABEL_PATTERN = re.compile(
    r"^(?:contact|kontakt)(?P<index>\d+)(?P<field>role|rolle|name|email|phone|telefon)$"
)
_META_PREFIX_PATTERN = re.compile(r"^meta[\s_.-]*", re.IGNORECASE)


class SpiritLookupApp:
    """Tkinter based UI application."""

    def __init__(self, root: tk.Tk, controller: SpiritLookupController, config: AppConfig) -> None:
        self.root = root
        self.controller = controller
        self.config = config
        self.helper_config_path = config.fixture_path.parent / "excel_helper_config.json"
        self.helper_config_store = ExcelHelperConfigStore(self.helper_config_path)
        self.display_config_path = config.fixture_path.parent / "display_config.json"
        self.display_config = DisplayConfig(self.display_config_path)
        self.display_definitions: list[DisplayFieldDefinition] = []

        self.current_query: str = ""
        self.current_page: int = 0
        self.cached_records: List[SpiritRecord] = []
        self.has_more: bool = False
        self._debounce_id: str | None = None

        self.status_var = tk.StringVar(value="Bereit.")
        self.search_var = tk.StringVar()
        self.spirit_entry_var = tk.StringVar()

        self.setup_excel_path: Path | None = None
        self.setup_sheet_var = tk.StringVar()
        self.setup_warning_var = tk.StringVar()
        self.setup_records: list[dict[str, object]] = []
        self._excel_tool_module: ModuleType | None = None
        self._email_check_vars: list[tk.BooleanVar] = []

        self._load_display_config()
        self._build_ui()
        self._restore_excel_selection()
        self._load_initial()

    def _build_ui(self) -> None:
        self.root.title("Spirit Lookup")
        self.root.geometry("900x520")
        self.root.minsize(760, 420)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

        container = ttk.Frame(self.root)
        container.grid(row=0, column=0, sticky="nsew")
        container.columnconfigure(0, weight=1)
        container.rowconfigure(0, weight=1)

        notebook = ttk.Notebook(container)
        notebook.grid(row=0, column=0, sticky="nsew")

        # --- Hotel search tab ---
        search_frame = ttk.Frame(notebook, padding="16")
        notebook.add(search_frame, text="HotelSuche")

        for col in range(3):
            search_frame.columnconfigure(col, weight=1 if col < 2 else 0)

        ttk.Label(search_frame, text="Spirit Code eingeben:").grid(row=0, column=0, sticky="w")
        self.spirit_entry = ttk.Entry(search_frame, textvariable=self.spirit_entry_var, width=30)
        self.spirit_entry.grid(row=1, column=0, sticky="we", padx=(0, 12))
        self.spirit_entry.bind("<Return>", lambda _event: self.on_search())

        ttk.Label(search_frame, text="oder Hotel auswählen:").grid(row=0, column=1, sticky="w")
        self.search_combo = ttk.Combobox(search_frame, textvariable=self.search_var, width=50)
        self.search_combo.grid(row=1, column=1, sticky="we")
        self.search_combo.bind("<KeyRelease>", self._on_search_var_change)
        self.search_combo.bind("<<ComboboxSelected>>", lambda _event: self.on_search())

        search_button = ttk.Button(search_frame, text="Suchen", command=self.on_search)
        search_button.grid(row=1, column=2, padx=(12, 0))

        self.load_more_button = ttk.Button(
            search_frame,
            text="Weitere Ergebnisse laden",
            command=self._load_next_page,
        )
        self.load_more_button.grid(row=2, column=1, sticky="e", pady=(8, 0))
        self.load_more_button.grid_remove()

        # --- Setup tab ---
        setup_frame = ttk.Frame(notebook, padding="16")
        notebook.add(setup_frame, text="Setup")
        setup_frame.columnconfigure(0, weight=1)
        setup_frame.rowconfigure(8, weight=1)

        ttk.Label(
            setup_frame,
            text=(
                "Verwalten Sie hier die Excel-Konfiguration und konvertieren Sie die Daten direkt in das "
                "Fixture-JSON."
            ),
            wraplength=620,
        ).grid(row=0, column=0, sticky="w")

        helper_button = ttk.Button(setup_frame, text="Excel Helper öffnen", command=self._open_excel_helper)
        helper_button.grid(row=1, column=0, sticky="w", pady=(12, 0))

        ttk.Label(
            setup_frame,
            text=f"Konfigurationsdatei: {self.helper_config_path}",
            foreground="#555555",
        ).grid(row=2, column=0, sticky="w", pady=(4, 12))

        ttk.Separator(setup_frame).grid(row=3, column=0, sticky="we", pady=(0, 12))

        choose_frame = ttk.Frame(setup_frame)
        choose_frame.grid(row=4, column=0, sticky="we")
        choose_frame.columnconfigure(1, weight=1)

        ttk.Button(choose_frame, text="Excel-Datei wählen", command=self._setup_choose_excel).grid(
            row=0, column=0, sticky="w"
        )
        self.setup_file_label = ttk.Label(
            choose_frame,
            text="Keine Datei gewählt",
            foreground="#555555",
        )
        self.setup_file_label.grid(row=0, column=1, sticky="w", padx=(12, 0))

        sheet_frame = ttk.Frame(setup_frame)
        sheet_frame.grid(row=5, column=0, sticky="we", pady=(12, 0))
        sheet_frame.columnconfigure(1, weight=1)
        ttk.Label(sheet_frame, text="Arbeitsblatt:").grid(row=0, column=0, sticky="w")
        self.setup_sheet_combo = ttk.Combobox(
            sheet_frame,
            textvariable=self.setup_sheet_var,
            state="disabled",
        )
        self.setup_sheet_combo.grid(row=0, column=1, sticky="we")

        action_frame = ttk.Frame(setup_frame)
        action_frame.grid(row=6, column=0, sticky="we", pady=(12, 0))
        action_frame.columnconfigure(3, weight=1)
        self.setup_convert_button = ttk.Button(
            action_frame,
            text="Excel einlesen",
            command=self._setup_convert_excel,
            state=tk.DISABLED,
        )
        self.setup_convert_button.grid(row=0, column=0, sticky="w")
        self.setup_save_button = ttk.Button(
            action_frame,
            text="JSON speichern",
            command=self._setup_save_json,
            state=tk.DISABLED,
        )
        self.setup_save_button.grid(row=0, column=1, sticky="w", padx=(12, 0))
        self.setup_generate_display_button = ttk.Button(
            action_frame,
            text="Anzeige-JSON speichern",
            command=self._setup_generate_display_config,
            state=tk.DISABLED,
        )
        self.setup_generate_display_button.grid(row=0, column=2, sticky="w", padx=(12, 0))

        self.setup_warning_label = ttk.Label(
            setup_frame,
            textvariable=self.setup_warning_var,
            foreground="#a46400",
            wraplength=620,
        )
        self.setup_warning_label.grid(row=7, column=0, sticky="we")

        preview_frame = ttk.Frame(setup_frame)
        preview_frame.grid(row=8, column=0, sticky="nsew", pady=(12, 0))
        preview_frame.columnconfigure(0, weight=1)
        preview_frame.rowconfigure(0, weight=1)
        self.setup_preview = tk.Text(preview_frame, wrap="none", height=12)
        self.setup_preview.grid(row=0, column=0, sticky="nsew")
        preview_scroll = ttk.Scrollbar(preview_frame, orient="vertical", command=self.setup_preview.yview)
        preview_scroll.grid(row=0, column=1, sticky="ns")
        self.setup_preview.configure(yscrollcommand=preview_scroll.set, state="disabled")

        ttk.Label(
            setup_frame,
            text="Hinweis: Gespeicherte Dateien werden standardmäßig im data/ Verzeichnis abgelegt.",
            foreground="#555555",
        ).grid(row=9, column=0, sticky="w", pady=(12, 0))

        self._set_setup_preview("[]")
        self.setup_warning_var.set("")

        self.status_label = ttk.Label(container, textvariable=self.status_var, relief=tk.SUNKEN, anchor="w")
        self.status_label.grid(row=1, column=0, sticky="we", padx=16, pady=(8, 16))

    def _update_status(self, text: str) -> None:
        self.status_var.set(text)

    def _load_display_config(self) -> None:
        self.display_config.load()
        self.display_definitions = list(self.display_config.fields)

    def _update_display_button_state(self) -> None:
        if not hasattr(self, "setup_generate_display_button"):
            return
        entry_path = self.setup_excel_path or self.helper_config_store.get_last_used_path()
        entry = self.helper_config_store.get_entry(entry_path) if entry_path else None
        state = tk.NORMAL if entry and entry.selected_columns else tk.DISABLED
        self.setup_generate_display_button.configure(state=state)

    def _open_excel_helper(self) -> None:
        open_excel_helper(self.root, self.helper_config_path)
        self.helper_config_store.reload()
        self._restore_excel_selection(auto_convert=True)

    def _restore_excel_selection(self, *, auto_convert: bool = False) -> None:
        last_used = self.helper_config_store.get_last_used_path()
        if not last_used:
            self.setup_excel_path = None
            self._update_display_button_state()
            return
        entry = self.helper_config_store.get_entry(last_used)
        if not entry:
            self.setup_excel_path = None
            self._update_display_button_state()
            return
        if not last_used.exists():
            self.setup_excel_path = None
            self.setup_file_label.configure(text=f"{last_used} (nicht gefunden)")
            self.setup_sheet_combo.configure(state="disabled")
            self.setup_sheet_combo["values"] = []
            self.setup_sheet_var.set("")
            self.setup_convert_button.configure(state=tk.DISABLED)
            self.setup_save_button.configure(state=tk.DISABLED)
            pretty = self.helper_config_store.to_pretty_json(last_used)
            self._set_setup_preview(pretty)
            self.setup_warning_var.set(
                "Die konfigurierte Excel-Datei wurde nicht gefunden. Bitte wählen Sie eine Datei aus."
            )
            self._update_display_button_state()
            return
        try:
            sheet_names = self._load_sheet_names(last_used)
        except ModuleNotFoundError as exc:
            self.setup_warning_var.set(str(exc))
            self.setup_convert_button.configure(state=tk.DISABLED)
            self.setup_save_button.configure(state=tk.DISABLED)
            pretty = self.helper_config_store.to_pretty_json(last_used)
            self._set_setup_preview(pretty if pretty != "{}" else "[]")
            self._update_display_button_state()
            return
        except ValueError as exc:
            self.setup_warning_var.set(str(exc))
            self.setup_convert_button.configure(state=tk.DISABLED)
            self.setup_save_button.configure(state=tk.DISABLED)
            pretty = self.helper_config_store.to_pretty_json(last_used)
            self._set_setup_preview(pretty if pretty != "{}" else "[]")
            self._update_display_button_state()
            return

        self.setup_excel_path = last_used
        self.setup_file_label.configure(text=str(last_used))
        self.setup_sheet_combo.configure(state="readonly")
        self.setup_sheet_combo["values"] = sheet_names
        if sheet_names:
            self.setup_sheet_var.set(sheet_names[0])
        self.setup_convert_button.configure(state=tk.NORMAL)
        self.setup_save_button.configure(state=tk.DISABLED)
        self.setup_records = []
        pretty_config = self.helper_config_store.to_pretty_json(last_used)
        if pretty_config != "{}":
            self._set_setup_preview(pretty_config)
        else:
            self._set_setup_preview("[]")
        self.setup_warning_var.set("Konfiguration geladen. Bitte Excel einlesen.")
        self._update_display_button_state()
        if auto_convert and sheet_names:
            self.root.after(50, self._setup_convert_excel)

    def _set_setup_preview(self, content: str) -> None:
        self.setup_preview.configure(state="normal")
        self.setup_preview.delete("1.0", tk.END)
        self.setup_preview.insert("1.0", content)
        self.setup_preview.configure(state="disabled")

    def _get_excel_tool_module(self) -> ModuleType | None:
        if self._excel_tool_module is not None:
            return self._excel_tool_module
        try:
            from tools import excel_to_fixture as excel_tool  # type: ignore
        except ModuleNotFoundError as exc:
            messagebox.showerror("Excel-Import nicht möglich", str(exc))
            return None
        self._excel_tool_module = excel_tool
        return excel_tool

    def _load_sheet_names(self, excel_path: Path) -> List[str]:
        try:
            from openpyxl import load_workbook  # type: ignore
        except ModuleNotFoundError as exc:  # pragma: no cover - optional dependency
            raise ModuleNotFoundError(
                "Für den Excel-Import wird 'openpyxl' benötigt. Installieren Sie das Paket über `pip install openpyxl`."
            ) from exc
        try:
            workbook = load_workbook(excel_path, data_only=True, read_only=True)
        except Exception as exc:
            raise ValueError(f"Die Excel-Datei '{excel_path}' konnte nicht geöffnet werden: {exc}") from exc
        sheet_names = list(workbook.sheetnames)
        workbook.close()
        if not sheet_names:
            raise ValueError("Die Arbeitsmappe enthält keine Arbeitsblätter.")
        return sheet_names

    def _setup_choose_excel(self) -> None:
        file_path = filedialog.askopenfilename(
            parent=self.root,
            title="Excel-Datei auswählen",
            filetypes=[("Excel-Dateien", "*.xlsx"), ("Alle Dateien", "*.*")],
        )
        if not file_path:
            return
        excel_path = Path(file_path)
        if not excel_path.exists():
            messagebox.showerror("Datei nicht gefunden", f"Die Datei '{excel_path}' wurde nicht gefunden.")
            return
        try:
            sheet_names = self._load_sheet_names(excel_path)
        except ModuleNotFoundError as exc:
            messagebox.showerror("openpyxl nicht installiert", str(exc))
            return
        except ValueError as exc:
            messagebox.showerror("Fehler beim Einlesen", str(exc))
            return

        self.setup_excel_path = excel_path
        self.setup_file_label.configure(text=str(excel_path))
        self.setup_sheet_combo.configure(state="readonly")
        self.setup_sheet_combo["values"] = sheet_names
        self.setup_sheet_var.set(sheet_names[0])
        self.setup_convert_button.configure(state=tk.NORMAL)
        self.setup_save_button.configure(state=tk.DISABLED)
        self.setup_records = []
        self.setup_warning_var.set("Noch keine Konvertierung durchgeführt.")
        self._set_setup_preview("[]")
        self._update_display_button_state()

    def _setup_convert_excel(self) -> None:
        if not self.setup_excel_path:
            messagebox.showinfo("Keine Datei", "Bitte wählen Sie zuerst eine Excel-Datei aus.")
            return
        excel_tool = self._get_excel_tool_module()
        if excel_tool is None:
            return
        sheet_name_value = self.setup_sheet_var.get().strip() or None
        try:
            records, warnings = excel_tool.convert_excel_to_fixture(
                self.setup_excel_path,
                sheet_name=sheet_name_value,
            )
        except ValueError as exc:
            messagebox.showerror("Konvertierung fehlgeschlagen", str(exc))
            self.setup_records = []
            self.setup_save_button.configure(state=tk.DISABLED)
            self.setup_warning_var.set("Konvertierung fehlgeschlagen.")
            self._set_setup_preview("[]")
            return

        self.setup_records = list(records)
        pretty = json.dumps(self.setup_records, indent=2, ensure_ascii=False)
        self._set_setup_preview(pretty)
        if warnings:
            warning_text = "Warnungen:\n" + "\n".join(f"• {warning}" for warning in warnings)
            self.setup_warning_var.set(warning_text)
        else:
            self.setup_warning_var.set("Keine Warnungen.")
        self.setup_save_button.configure(state=tk.NORMAL)
        self._update_status(f"{len(self.setup_records)} Datensätze vorbereitet.")

    def _setup_save_json(self) -> None:
        if not self.setup_records:
            messagebox.showinfo("Keine Daten", "Bitte lesen Sie zunächst eine Excel-Datei ein.")
            return
        excel_tool = self._get_excel_tool_module()
        if excel_tool is None:
            return
        initialdir = str(self.config.fixture_path.parent)
        initialfile = self.config.fixture_path.name
        file_path = filedialog.asksaveasfilename(
            parent=self.root,
            title="JSON speichern",
            defaultextension=".json",
            filetypes=[("JSON-Dateien", "*.json"), ("Alle Dateien", "*.*")],
            initialdir=initialdir,
            initialfile=initialfile,
        )
        if not file_path:
            return
        output_path = Path(file_path)
        try:
            excel_tool.write_fixture(self.setup_records, output_path, indent=2)
        except OSError as exc:
            messagebox.showerror("Speichern fehlgeschlagen", f"Die Datei konnte nicht gespeichert werden: {exc}")
            return
        self._update_status(f"JSON nach '{output_path}' gespeichert.")
        messagebox.showinfo("Gespeichert", f"Die JSON-Datei wurde unter '{output_path}' gespeichert.")

    def _setup_generate_display_config(self) -> None:
        entry_path = self.setup_excel_path or self.helper_config_store.get_last_used_path()
        if not entry_path:
            messagebox.showinfo(
                "Keine Konfiguration",
                "Es ist keine gespeicherte Excel-Konfiguration vorhanden. Bitte wählen Sie zunächst eine Datei aus.",
            )
            return
        entry = self.helper_config_store.get_entry(entry_path)
        if not entry or not entry.selected_columns:
            messagebox.showinfo(
                "Keine Spalten",
                "Es wurden keine Spalten gespeichert. Bitte nutzen Sie zuerst den Excel Helper, um Spalten auszuwählen.",
            )
            self._update_display_button_state()
            return

        definitions = [
            DisplayFieldDefinition(label=column, is_email=column in entry.email_columns)
            for column in entry.selected_columns
        ]
        try:
            self.display_config.save(definitions)
        except OSError as exc:
            messagebox.showerror("Speichern fehlgeschlagen", f"Die Anzeige-Konfiguration konnte nicht gespeichert werden: {exc}")
            return

        self.display_definitions = list(self.display_config.fields)
        self._update_status(f"Anzeige-Konfiguration nach '{self.display_config_path}' gespeichert.")
        messagebox.showinfo(
            "Anzeige-JSON gespeichert",
            f"Die Anzeige-Konfiguration wurde unter '{self.display_config_path}' gespeichert.",
        )

    def _default_display_definitions(self, record: SpiritRecord) -> List[DisplayFieldDefinition]:
        definitions = [
            DisplayFieldDefinition("Spirit Code"),
            DisplayFieldDefinition("Hotel Name"),
            DisplayFieldDefinition("Region"),
            DisplayFieldDefinition("Status"),
            DisplayFieldDefinition("City"),
            DisplayFieldDefinition("Country"),
            DisplayFieldDefinition("Address"),
        ]
        for index, _contact in enumerate(record.contacts, start=1):
            prefix = f"Kontakt {index}"
            definitions.extend(
                [
                    DisplayFieldDefinition(f"{prefix} Role"),
                    DisplayFieldDefinition(f"{prefix} Name"),
                    DisplayFieldDefinition(f"{prefix} Email", is_email=True),
                    DisplayFieldDefinition(f"{prefix} Phone"),
                ]
            )
        for key in sorted(record.meta.keys()):
            definitions.append(DisplayFieldDefinition(key))
        return definitions

    def _build_display_fields(self, record: SpiritRecord) -> List[RenderedField]:
        definitions = self.display_definitions or self._default_display_definitions(record)
        fields: List[RenderedField] = []
        for definition in definitions:
            raw_value = self._resolve_record_value(record, definition.label)
            formatted = self._format_field_value(raw_value)
            fields.append(RenderedField(label=definition.label, value=formatted, is_email=definition.is_email))
        return fields

    def _resolve_record_value(self, record: SpiritRecord, label: str) -> object | None:
        normalized = normalize_header(label)
        if not normalized:
            return None
        base_map = {
            "spiritcode": record.spirit_code,
            "hotelname": record.hotel_name,
            "hotel": record.hotel_name,
            "region": record.region,
            "status": record.status,
            "city": record.location_city,
            "locationcity": record.location_city,
            "country": record.location_country,
            "locationcountry": record.location_country,
            "address": record.address,
            "locationaddress": record.address,
        }
        if normalized in base_map:
            return base_map[normalized]

        contact_match = _CONTACT_LABEL_PATTERN.match(normalized)
        if contact_match:
            index = int(contact_match.group("index")) - 1
            field = contact_match.group("field")
            field_map = {
                "role": "role",
                "rolle": "role",
                "name": "name",
                "email": "email",
                "phone": "phone",
                "telefon": "phone",
            }
            attr = field_map.get(field)
            if attr and 0 <= index < len(record.contacts):
                return getattr(record.contacts[index], attr)

        meta_value = self._lookup_meta_value(record, label, normalized)
        if meta_value is not None:
            return meta_value

        return None

    def _lookup_meta_value(self, record: SpiritRecord, label: str, normalized: str) -> object | None:
        if not record.meta:
            return None
        meta_key = self._derive_meta_key(label)
        if meta_key in record.meta:
            return record.meta[meta_key]
        for key, value in record.meta.items():
            if normalize_header(key) == normalized:
                return value
        if label in record.meta:
            return record.meta[label]
        return None

    def _derive_meta_key(self, label: str) -> str:
        stripped = _META_PREFIX_PATTERN.sub("", label.strip())
        if not stripped:
            stripped = "metaField"
        parts = [part for part in re.split(r"[^0-9A-Za-z]+", stripped) if part]
        if not parts:
            return "metaField"
        first, *rest = parts
        camel = first[:1].lower() + first[1:]
        for piece in rest:
            camel += piece[:1].upper() + piece[1:]
        return camel

    def _format_field_value(self, value: object | None) -> str | None:
        if value is None:
            return None
        if isinstance(value, bool):
            return "Ja" if value else "Nein"
        return str(value)

    def _load_initial(self) -> None:
        try:
            self._update_status("Lade Daten …")
            result = self.controller.list_records(page=0)
            self._apply_lookup_result(result)
            if not result.records:
                self._update_status("Keine Daten verfügbar. Nutzen Sie ggf. den Fixture-Modus.")
            else:
                self._update_status(f"{len(self.cached_records)} Hotels geladen.")
        except DataProviderError as exc:
            messagebox.showerror("Fehler beim Laden", str(exc))
            self._update_status("Laden fehlgeschlagen – bitte erneut versuchen.")

    def _apply_lookup_result(self, result: LookupResult) -> None:
        self.current_page = result.page
        self.has_more = result.has_more
        if result.page == 0:
            self.cached_records = list(result.records)
        else:
            known_codes = {record.spirit_code for record in self.cached_records}
            for record in result.records:
                if record.spirit_code not in known_codes:
                    self.cached_records.append(record)
                    known_codes.add(record.spirit_code)

        values = [record.display_label() for record in self.cached_records]
        if not values:
            self.search_combo["values"] = [""]
        else:
            self.search_combo["values"] = values
        if self.has_more:
            self.load_more_button.grid()
            self.load_more_button.configure(state=tk.NORMAL)
        else:
            self.load_more_button.grid_remove()

    def _schedule_search(self, query: str) -> None:
        if self._debounce_id:
            self.root.after_cancel(self._debounce_id)
        self._debounce_id = self.root.after(self.config.debounce_ms, lambda: self._perform_search(query))

    def _on_search_var_change(self, _event) -> None:
        query = self.search_var.get().strip()
        self.current_query = query
        self._schedule_search(query)

    def _perform_search(self, query: str, page: int = 0) -> None:
        try:
            self._update_status("Suche …")
            result = self.controller.list_records(query, page=page)
            self._apply_lookup_result(result)
            if not result.records:
                self._update_status("Keine Treffer – versuchen Sie eine andere Suche.")
            else:
                self._update_status(
                    f"{len(result.records)} Treffer auf Seite {result.page + 1}."
                    + (" Weitere Seiten verfügbar." if result.has_more else "")
                )
        except DataProviderError as exc:
            retry = messagebox.askretrycancel("Fehler", f"{exc}\nErneut versuchen?")
            if retry:
                self._perform_search(query, page=page)
            else:
                self._update_status("Suche abgebrochen.")

    def _load_next_page(self) -> None:
        if not self.has_more:
            return
        next_page = self.current_page + 1
        self._perform_search(self.current_query, page=next_page)

    def on_search(self) -> None:
        try:
            record = self.controller.search_by_input(
                spirit_code=self.spirit_entry_var.get().strip() or None,
                selected_label=self.search_var.get().strip() or None,
                cached_records=self.cached_records,
            )
            self._show_details(record)
        except RecordNotFoundError as exc:
            messagebox.showinfo("Keine Auswahl", str(exc))
        except DataProviderError as exc:
            retry = messagebox.askretrycancel("Fehler", f"{exc}\nErneut versuchen?")
            if retry:
                self.on_search()
        finally:
            self.spirit_entry.selection_clear()

    def _show_details(self, record: SpiritRecord) -> None:
        dialog = tk.Toplevel(self.root)
        dialog.title(f"{record.hotel_name} ({record.spirit_code})")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.geometry("640x480")
        dialog.resizable(True, True)

        dialog.columnconfigure(0, weight=1)
        dialog.rowconfigure(1, weight=1)

        ttk.Label(dialog, text="Schlüssel-Informationen", font=("Segoe UI", 12, "bold")).grid(
            row=0, column=0, sticky="nw", padx=16, pady=(16, 4)
        )

        info_frame = ttk.Frame(dialog, padding="16")
        info_frame.grid(row=1, column=0, sticky="nsew")
        info_frame.columnconfigure(1, weight=1)
        info_frame.columnconfigure(3, weight=0)

        fields = self._build_display_fields(record)
        self._email_check_vars = []
        if not fields:
            ttk.Label(info_frame, text="Keine Felder definiert.").grid(row=0, column=0, sticky="w")
        else:
            for idx, field in enumerate(fields):
                label_widget = ttk.Label(
                    info_frame,
                    text=f"{field.label}:",
                    font=("Segoe UI", 10, "bold"),
                    anchor="w",
                )
                label_widget.grid(row=idx, column=0, sticky="w", pady=2)
                value_text = field.value if field.value not in (None, "") else "–"
                value_label = ttk.Label(info_frame, text=value_text, anchor="w", wraplength=360)
                value_label.grid(row=idx, column=1, sticky="w", pady=2)

                column_offset = 2
                if field.is_email and field.value:
                    var = tk.BooleanVar(value=False)
                    self._email_check_vars.append(var)
                    ttk.Checkbutton(info_frame, variable=var).grid(row=idx, column=2, padx=(8, 0), sticky="w")
                    column_offset = 3
                if field.value:
                    ttk.Button(
                        info_frame,
                        text="Kopieren",
                        command=lambda value=field.value: self._copy_to_clipboard(value),
                    ).grid(row=idx, column=column_offset, padx=(8, 0), sticky="w")

        action_frame = ttk.Frame(dialog, padding="16")
        action_frame.grid(row=2, column=0, sticky="ew")
        action_frame.columnconfigure(2, weight=1)

        confirm_var = tk.BooleanVar(value=False)
        confirm_check = ttk.Checkbutton(
            action_frame,
            text="Hinweise gelesen / Datenschutz akzeptiert",
            variable=confirm_var,
            command=lambda: draft_button.configure(state=(tk.NORMAL if confirm_var.get() else tk.DISABLED)),
        )
        confirm_check.grid(row=0, column=0, sticky="w")

        draft_button = ttk.Button(
            action_frame,
            text="Draft E-Mail",
            command=lambda: self._open_draft_email(dialog),
            state=tk.NORMAL if (self.config.draft_email_enabled and confirm_var.get()) else tk.DISABLED,
        )
        draft_button.grid(row=0, column=1, padx=(12, 0))

        if not self.config.draft_email_enabled:
            draft_button.configure(state=tk.DISABLED)

        close_button = ttk.Button(action_frame, text="Schließen", command=lambda: self._close_dialog(dialog))
        close_button.grid(row=0, column=2, sticky="e")

        def on_confirm_change(*_args):
            if self.config.draft_email_enabled:
                draft_button.configure(state=tk.NORMAL if confirm_var.get() else tk.DISABLED)

        confirm_var.trace_add("write", on_confirm_change)

        dialog.bind("<Escape>", lambda _event: self._close_dialog(dialog))
        dialog.focus_set()
        close_button.focus_set()

    def _copy_to_clipboard(self, value: str) -> None:
        self.root.clipboard_clear()
        self.root.clipboard_append(value)
        self._update_status(f"'{value}' in Zwischenablage kopiert.")

    def _close_dialog(self, dialog: tk.Toplevel) -> None:
        dialog.grab_release()
        dialog.destroy()
        self.spirit_entry.focus_set()

    def _open_draft_email(self, dialog: tk.Toplevel) -> None:
        try:
            open_mail_client("mailto:")
        except MailClientError as exc:
            messagebox.showerror("E-Mail", str(exc))
        else:
            self._update_status("Mailclient wurde geöffnet.")
        finally:
            dialog.grab_release()
            dialog.destroy()
            self.spirit_entry.focus_set()


def run_app(config: AppConfig, controller: SpiritLookupController) -> None:
    root = tk.Tk()
    app = SpiritLookupApp(root, controller, config)
    root.mainloop()
