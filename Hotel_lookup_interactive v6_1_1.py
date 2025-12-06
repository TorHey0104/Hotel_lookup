#!/usr/bin/env python3
import os
import shutil
from datetime import datetime

# Ensure pandas (and openpyxl) are available; try to import and attempt to install if missing
# Add "# type: ignore" to silence editors/linters that cannot resolve the package in the current environment.
try:
    import pandas as pd  # type: ignore
except Exception:
    import sys
    import subprocess
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "--user", "pandas", "openpyxl"])
        import importlib
        importlib.invalidate_caches()
        import pandas as pd  # type: ignore
    except Exception as e:  # pragma: no cover - auto-install fallback
        raise ImportError(
            "Could not import or install 'pandas' and/or 'openpyxl'. Please install them manually (e.g. pip install pandas openpyxl)."
        ) from e

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
import re
import importlib.util
import glob
import openpyxl
import html
import math
import tempfile
from datetime import date
from typing import Optional
from openpyxl.utils import get_column_letter, range_boundaries

# Cache Outlook availability and instance so email drafting is faster after the first use
WIN32COM_AVAILABLE = os.name == "nt" and importlib.util.find_spec("win32com.client") is not None
_outlook_app = None

# ---------------------------------------------------------------------------
# CONFIGURE THIS
# ---------------------------------------------------------------------------
DATA_DIR = r"C:\Users\4612135\OneDrive - Hyatt Hotels\___DATA"
FILE_NAME = "2a Hotels one line hotel.xlsx"
DEFAULT_FILE_PATH = os.path.join(DATA_DIR, FILE_NAME)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LOGO_PATH = os.path.join(BASE_DIR, "hyatt_logo.png")  # optional logo next to script
RECENT_CONFIG_PATH = os.path.join(BASE_DIR, "recent_configs.json")

TOOL_NAME = "Hyatt EAME Hotel Lookup and Multi E-Mail Tool"
VERSION = "6.1.1"
VERSION_DATE = date.today().strftime("%d.%m.%Y")

# Default column names (can be overridden in Setup tab)
DEFAULT_BRAND_COL = "Brand"
DEFAULT_REGION_COL = "Region"
DEFAULT_COUNTRY_COL = "Geography"
DEFAULT_COUNTRY_FALLBACK_COL = "Geographical Area"
DEFAULT_CITY_COL = "City"
DEFAULT_BRAND_BAND_COL = "Brand Band"
DEFAULT_RELATIONSHIP_COL = "Relationship"
DEFAULT_HYATT_DATE_COL = "Affiliation Date"
DEFAULT_GM_COL = "GM - Primary"
DEFAULT_ENG_COL = "Engineering Director / Chief Engineer"
DEFAULT_DOF_COL = "DOF"
DEFAULT_REG_ENG_SPEC_COL = ""  # optional

# Runtime data containers (populated by load_data)
df = pd.DataFrame()
hotel_names = []
data_file_path = ""
brand_values = []
region_values = []
country_values = []
brand_band_values = []
relationship_values = []

# Tk widgets / state
hotel_combo = None
status_var = None
brand_filter_var = None  # kept for legacy; not used with multiselect
region_filter_var = None  # kept for legacy; not used with multiselect
country_filter_var = None  # kept for legacy; not used with multiselect
brand_band_filter_var = None  # kept for legacy; not used with multiselect
relationship_filter_var = None  # kept for legacy; not used with multiselect
hyatt_year_var = None
hyatt_year_mode_var = None
brand_listbox = None
brand_band_listbox = None
region_listbox = None
relationship_listbox = None
country_listbox = None
filtered_tree = None
selected_tree = None
selected_rows = {}
current_filtered_indexes = []
role_send_vars = {}
ROLE_MODES = ["Skip", "To", "CC", "BCC"]
style = None
filtered_count_var = None

# Column selection vars (set in setup tab)
brand_col_var = None
region_col_var = None
country_col_var = None
country_fallback_col_var = None
city_col_var = None
gm_col_var = None
eng_col_var = None
dof_col_var = None
reg_eng_spec_col_var = None
avp_col_var = None
md_col_var = None
brand_band_col_var = None
hyatt_date_col_var = None
relationship_col_var = None

brand_col_combo = None
region_col_combo = None
country_col_combo = None
country_fallback_combo = None
city_col_combo = None
gm_col_combo = None
eng_col_combo = None
dof_col_combo = None
reg_eng_spec_combo = None
avp_col_combo = None
md_col_combo = None
brand_band_col_combo = None
hyatt_date_col_combo = None
relationship_col_combo = None

# Lookup detail panel state
detail_info_vars = {}
detail_roles_frame = None
detail_checkbox_vars = []
detail_hotel_name = ""
detail_status_var = None
detail_start_email_btn = None
detail_row_current = None
signatures_cache = {}
splash_win = None
splash_status_var = None
splash_file_var = None
splash_logo_img = None
config_prompted = False
# Single compose UI state
single_compose_frame = None
single_subj_var = None
single_body_text = None
single_sig_var = None
single_recipient_controls = []
single_recips_frame = None
# Attachments
attachments_enabled_var = None
attachments_root_var = None
attachments_common_dir = "Common"
attachments_spirit_dir = "Spirit"
single_attachments_enabled_var = None
single_attachments_root_var = None
forward_template = {"subject": "", "body_text": "", "attachments": [], "temp_dir": ""}
forward_status_var = None
forward_template["is_html"] = False
quick_spirit_var = None
excel_df = pd.DataFrame()
excel_mapping_controls = []
excel_mapping = {}
excel_headers_frame = None
excel_file_label_var = None
excel_mode_var = None
excel_filter_controls = []
excel_filter_summary_var = None
excel_filtered_tree = None
excel_selected_tree = None
excel_filtered_cache = []
excel_selected_cache = {}

PLACEHOLDERS = [
    "{hotel}",
    "{spirit_code}",
    "{city}",
    "{relationship}",
    "{brand}",
    "{brand_band}",
    "{region}",
    "{country}",
    "{owner}",
    "{rooms}",
]

# Visible columns for filtered hotels
MANDATORY_FILTER_COLS = ["Spirit Code", "Hotel"]
visible_optional_filter_cols = ["City", "Brand", "Brand Band", "Relationship", "Region", "Country"]
filter_cols_listbox = None


def format_timestamp(path: str) -> str:
    """Return a human friendly timestamp for the given file path."""
    try:
        mod_time = datetime.fromtimestamp(os.path.getmtime(path))
    except (FileNotFoundError, OSError):
        return "Unknown timestamp"
    return mod_time.strftime("%d.%m.%Y %H:%M")


def get_selected_col(var: tk.StringVar | None, allow_none: bool = False) -> str:
    if var is None:
        return ""
    val = var.get().strip()
    if val == "None":
        return ""
    return val


def get_brand_col():
    return get_selected_col(brand_col_var)


def get_region_col():
    return get_selected_col(region_col_var)


def get_city_col():
    return get_selected_col(city_col_var)


def get_country_col():
    return get_selected_col(country_col_var)


def get_country_fallback_col():
    return get_selected_col(country_fallback_col_var, allow_none=True)


def get_brand_band_col():
    return get_selected_col(brand_band_col_var, allow_none=True)

def get_relationship_col():
    return get_selected_col(relationship_col_var, allow_none=True)


def get_gm_col():
    return get_selected_col(gm_col_var)


def get_eng_col():
    return get_selected_col(eng_col_var)


def get_dof_col():
    return get_selected_col(dof_col_var)


def get_reg_eng_spec_col():
    return get_selected_col(reg_eng_spec_col_var, allow_none=True)


def get_avp_col():
    return get_selected_col(avp_col_var, allow_none=True)


def get_md_col():
    return get_selected_col(md_col_var, allow_none=True)


def get_hyatt_date_col():
    return get_selected_col(hyatt_date_col_var, allow_none=True)


def normalize_emails(raw: str):
    """Split a raw email string by common delimiters and drop placeholders like N/A."""
    parts = []
    for chunk in str(raw).replace(",", ";").split(";"):
        email = chunk.strip()
        if not email:
            continue
        low = email.lower()
        if low in {"n/a", "na", "none"}:
            continue
        parts.append(email)
    return parts


def render_template(row: pd.Series, template: str) -> str:
    """Replace placeholders in a template string using row values."""
    brand_col = get_brand_col()
    region_col = get_region_col()
    relationship_col = get_relationship_col()
    brand_band_col = get_brand_band_col()
    replacements = {
        "{hotel}": row.get("Hotel", ""),
        "{spirit_code}": row.get("Spirit Code", ""),
        "{city}": get_city_value(row),
        "{relationship}": row.get(relationship_col, "") if relationship_col in row else "",
        "{brand}": row.get(brand_col, "") if brand_col in row else "",
        "{brand_band}": row.get(brand_band_col, "") if brand_band_col in row else "",
        "{region}": row.get(region_col, "") if region_col in row else "",
        "{country}": get_country_value(row),
        "{owner}": row.get("Owner", ""),
        "{rooms}": row.get("Rooms", ""),
    }
    rendered = template
    for key, val in replacements.items():
        rendered = rendered.replace(key, str(val))
    return rendered


def collect_spirit_dirs(attach_root: str, spirit_code: str) -> list:
    """Return candidate directories for spirit-specific attachments."""
    if not spirit_code:
        return []
    candidates = [
        os.path.join(attach_root, attachments_spirit_dir, spirit_code),
        os.path.join(attach_root, spirit_code),
        os.path.join(attach_root, attachments_spirit_dir, spirit_code.upper()),
        os.path.join(attach_root, attachments_spirit_dir, spirit_code.lower()),
        os.path.join(attach_root, spirit_code.upper()),
        os.path.join(attach_root, spirit_code.lower()),
    ]
    uniq = []
    seen = set()
    for path in candidates:
        if path not in seen:
            seen.add(path)
            uniq.append(path)
    return uniq


def attach_files_for_hotel(mail_item, attach_root: str, spirit_code: str):
    """Attach common + spirit-specific files when available."""
    if not attach_root or not os.path.isdir(attach_root):
        return
    common_dir = os.path.join(attach_root, attachments_common_dir)
    if os.path.isdir(common_dir):
        for path in glob.glob(os.path.join(common_dir, "*")):
            if os.path.isfile(path):
                try:
                    mail_item.Attachments.Add(path)
                except Exception:
                    pass
    for candidate in collect_spirit_dirs(attach_root, spirit_code):
        if os.path.isdir(candidate):
            for path in glob.glob(os.path.join(candidate, "*")):
                if os.path.isfile(path):
                    try:
                        mail_item.Attachments.Add(path)
                    except Exception:
                        pass
            break


def browse_attachments_root():
    sel = filedialog.askdirectory(title="Choose attachments root")
    if sel:
        attachments_root_var.set(sel)


def capture_outlook_selection():
    """Deprecated: selection-based capture removed in v5.0.1."""
    messagebox.showinfo("Outlook", "This option has been removed. Please use ''Browse Outlook...'' instead.")

def clear_forward_template():
    """Clear cached forward email content/attachments."""
    # clean temp files
    if forward_template.get("temp_dir") and os.path.isdir(forward_template["temp_dir"]):
        try:
            for path in glob.glob(os.path.join(forward_template["temp_dir"], "*")):
                try:
                    os.remove(path)
                except Exception:
                    pass
        except Exception:
            pass
    forward_template["subject"] = ""
    forward_template["body_text"] = ""
    forward_template["is_html"] = False
    forward_template["attachments"] = []
    forward_template["temp_dir"] = ""
    if forward_status_var is not None:
        forward_status_var.set("No source email captured.")

def browse_outlook_email():
    """Simple subject search in Outlook Inbox to pick an email to forward."""
    if os.name != "nt" or not WIN32COM_AVAILABLE:
        messagebox.showerror("Outlook", "Outlook not available on this system.")
        return
    try:
        import win32com.client  # type: ignore
        try:
            outlook = get_outlook_app()
            ns = outlook.GetNamespace("MAPI")
            inbox = ns.GetDefaultFolder(6)  # Inbox
            sent = ns.GetDefaultFolder(5)   # Sent Items
        except Exception:
            outlook = get_outlook_app(force_refresh=True)
            ns = outlook.GetNamespace("MAPI")
            inbox = ns.GetDefaultFolder(6)
            sent = ns.GetDefaultFolder(5)
    except Exception as exc:
        messagebox.showerror("Outlook", f"Could not access Outlook folders: {exc}")
        return

    dlg = tk.Toplevel(root)
    dlg.title("Find Outlook Email")
    dlg.geometry("600x400")
    tk.Label(dlg, text="Subject contains:").pack(anchor="w", padx=8, pady=(8, 2))
    subj_var = tk.StringVar()
    tk.Entry(dlg, textvariable=subj_var).pack(fill="x", padx=8, pady=(0, 8))

    listbox = tk.Listbox(dlg, selectmode="browse")
    listbox.pack(fill="both", expand=True, padx=8, pady=4)
    scrollbar = ttk.Scrollbar(listbox, orient="vertical", command=listbox.yview)
    listbox.config(yscrollcommand=scrollbar.set)
    scrollbar.pack(side="right", fill="y")

    results = []

    def search():
        query = subj_var.get().lower()
        listbox.delete(0, tk.END)
        results.clear()
        for folder in [inbox, sent]:
            try:
                items = folder.Items
                items.Sort("[ReceivedTime]", True)
                count = 0
                for item in items:
                    if count >= 200:
                        break
                    subj = getattr(item, "Subject", "")
                    if query in subj.lower():
                        results.append(item)
                        sender = getattr(item, "SenderName", "") or getattr(item, "To", "")
                        listbox.insert(tk.END, f"{subj} | {sender}")
                    count += 1
            except Exception:
                continue

    def choose():
        sel = listbox.curselection()
        if not sel:
            messagebox.showinfo("Outlook", "Select an email from the list.")
            return
        item = results[sel[0]]
        dlg.destroy()
        try:
            forward_template["subject"] = f"FW: {getattr(item, 'Subject', '')}"
            html_body = getattr(item, "HTMLBody", "") or ""
            plain_body = getattr(item, "Body", "") or ""
            forward_template["body_text"] = html_body if html_body else plain_body
            forward_template["is_html"] = bool(html_body)
            # Clear prior temp
            if forward_template.get("temp_dir") and os.path.isdir(forward_template["temp_dir"]):
                for path in glob.glob(os.path.join(forward_template["temp_dir"], "*")):
                    try:
                        os.remove(path)
                    except Exception:
                        pass
            temp_dir = tempfile.mkdtemp(prefix="forward_src_")
            forward_template["temp_dir"] = temp_dir
            forward_template["attachments"] = []
            atts = getattr(item, "Attachments", None)
            if atts:
                for i in range(1, atts.Count + 1):
                    att = atts.Item(i)
                    try:
                        save_path = os.path.join(temp_dir, att.FileName)
                        att.SaveAsFile(save_path)
                        forward_template["attachments"].append(save_path)
                    except Exception:
                        pass
            msg = f"Captured email:\nSubject: {forward_template['subject']}\nAttachments: {len(forward_template['attachments'])}"
            if forward_status_var is not None:
                forward_status_var.set(msg.replace('\n', ' | '))
            messagebox.showinfo("Outlook", msg)
        except Exception as exc:
            messagebox.showerror("Outlook", f"Could not capture email: {exc}")

    tk.Button(dlg, text="Search", command=search).pack(side="left", padx=8, pady=4)
    tk.Button(dlg, text="Use Selected", command=choose).pack(side="right", padx=8, pady=4)
    search()


def load_signatures():
    """Load Outlook signature texts/html from the default signatures folder."""
    global signatures_cache
    if signatures_cache:
        return signatures_cache

    signatures_cache = {"None": {"html": "", "text": ""}}
    sig_dir = os.path.join(os.path.expandvars(r"%APPDATA%"), "Microsoft", "Signatures")
    if not os.path.isdir(sig_dir):
        return signatures_cache

    base_names = set()
    for ext in ("*.txt", "*.htm", "*.html"):
        for path in glob.glob(os.path.join(sig_dir, ext)):
            base_names.add(os.path.splitext(os.path.basename(path))[0])

    for name in base_names:
        txt_path = os.path.join(sig_dir, name + ".txt")
        htm_path = os.path.join(sig_dir, name + ".htm")
        html_path = os.path.join(sig_dir, name + ".html")
        sig_entry = {"html": "", "text": ""}

        if os.path.isfile(htm_path):
            try:
                with open(htm_path, "r", encoding="utf-8", errors="ignore") as fh:
                    sig_entry["html"] = fh.read()
            except Exception:
                sig_entry["html"] = ""
        elif os.path.isfile(html_path):
            try:
                with open(html_path, "r", encoding="utf-8", errors="ignore") as fh:
                    sig_entry["html"] = fh.read()
            except Exception:
                sig_entry["html"] = ""

        if os.path.isfile(txt_path):
            try:
                with open(txt_path, "r", encoding="utf-8", errors="ignore") as fh:
                    sig_entry["text"] = fh.read().strip()
            except Exception:
                sig_entry["text"] = ""

        signatures_cache[name] = sig_entry
    return signatures_cache


def render_with_signature(body_text: str, signature_entry: dict, body_is_html: bool = False, forward_html: str = "", forward_is_html: bool = False):
    """Return rendered content with signature before forwarded content, preserving line breaks."""
    sig_html = signature_entry.get("html", "") if isinstance(signature_entry, dict) else ""
    sig_txt = signature_entry.get("text", "") if isinstance(signature_entry, dict) else ""

    base_style = "font-family:'Aptos',sans-serif; font-size:12pt; line-height:1.4;"
    link_pattern = re.compile(r"\[([^\]]+)\]\(([^)]+)\)")
    anchor_pattern = re.compile(r"<a\s+[^>]*href=[\"']?([^>\"']+)[\"']?[^>]*>(.*?)</a>", re.IGNORECASE | re.DOTALL)

    def linkify_text(txt: str) -> str:
        """Convert markdown links and preserve existing anchors; escape other text and newlines."""
        parts = []
        idx = 0
        combined_pattern = re.compile(r"\[([^\]]+)\]\(([^)]+)\)|<a\s+[^>]*href=[\"']?([^>\"']+)[\"']?[^>]*>(.*?)</a>", re.IGNORECASE | re.DOTALL)
        for m in combined_pattern.finditer(txt):
            pre = txt[idx:m.start()]
            if pre:
                parts.append(html.escape(pre).replace("\n", "<br>"))
            label = ""
            url_raw = ""
            if m.group(1) is not None:
                label = m.group(1)
                url_raw = m.group(2)
            else:
                label = m.group(4)
                url_raw = m.group(3)
            url_raw = (url_raw or "").strip()
            if not url_raw.lower().startswith(("http://", "https://")):
                url_raw = "https://" + url_raw
            url = html.escape(url_raw)
            label_safe = html.escape(label or url_raw)
            parts.append(f'<a href="{url}">{label_safe}</a>')
            idx = m.end()
        tail = txt[idx:]
        if tail:
            parts.append(html.escape(tail).replace("\n", "<br>"))
        return "".join(parts) if parts else html.escape(txt).replace("\n", "<br>")

    def to_html(txt: str, allow_links: bool = False) -> str:
        return linkify_text(txt) if allow_links else html.escape(txt).replace("\n", "<br>")

    def looks_like_html(txt: str) -> bool:
        lowered = txt.lower()
        return any(tag in lowered for tag in ("<html", "<body", "<table", "<div", "<p"))

    def wrap_block(content: str) -> str:
        return f"<div style='white-space:pre-wrap; {base_style}'>{content}</div>"

    # User text to HTML block
    body_has_links = bool(link_pattern.search(body_text)) if not body_is_html else False
    if body_is_html:
        user_block = body_text  # already HTML
    else:
        user_block = wrap_block(to_html(body_text, allow_links=True))

    # Signature block
    sig_block = ""
    if sig_html:
        sig_block = wrap_block(sig_html if looks_like_html(sig_html) else to_html(sig_html, allow_links=True))
    elif sig_txt:
        sig_block = wrap_block(to_html(sig_txt, allow_links=True))

    # Forward block
    forward_block = ""
    if forward_html:
        if forward_is_html or looks_like_html(forward_html):
            forward_block = forward_html
        else:
            forward_block = wrap_block(to_html(forward_html, allow_links=True))

    # Build HTML if any rich content exists
    if forward_block or sig_block or body_is_html or forward_is_html or sig_html or body_has_links:
        html_parts = [user_block]
        if sig_block:
            html_parts.append(sig_block)
        if forward_block:
            html_parts.append(forward_block)
        html_body = "<br><br>".join([p for p in html_parts if p])
        if "<html" not in html_body.lower():
            html_body = f"<!DOCTYPE html><html><body style=\"{base_style}\">{html_body}</body></html>"
        return {"html": html_body}

    # Plain text fallback
    combined = body_text
    if sig_txt:
        combined += "\n\n" + sig_txt
    if forward_html:
        combined += "\n\n" + forward_html
    return {"text": combined}


def clear_single_compose():
    """Reset single-email compose fields."""
    global single_recipient_controls
    if single_subj_var is not None:
        single_subj_var.set("")
    if single_body_text is not None:
        single_body_text.delete("1.0", "end")
    if single_sig_var is not None:
        single_sig_var.set("None")
    if single_recips_frame is not None:
        for widget in single_recips_frame.winfo_children():
            widget.destroy()
    single_recipient_controls = []


def update_single_compose(row: pd.Series):
    """Populate single-email compose UI with the selected hotel."""
    if single_subj_var is None or single_body_text is None:
        return
    hotel_name = row.get("Hotel", "Hotel")
    single_subj_var.set(f"Hotel Information for {hotel_name}")
    body_template = "Hotel: {hotel}\nSpirit: {spirit_code}\nCity: {city}\nBrand: {brand}\n\nYour message here."
    single_body_text.delete("1.0", "end")
    single_body_text.insert("1.0", render_template(row, body_template))

    # Recipients
    if single_recips_frame is not None:
        for widget in single_recips_frame.winfo_children():
            widget.destroy()
    single_recipient_controls.clear()
    roles_to_col = {
        "AVP": get_avp_col(),
        "MD": get_md_col(),
        "GM": get_gm_col(),
        "Engineering": get_eng_col(),
        "DOF": get_dof_col(),
        "Regional Eng Specialist": get_reg_eng_spec_col(),
    }
    for role, col in roles_to_col.items():
        if not col:
            continue
        email = row.get(col)
        if col in row.index and pd.notna(email):
            sel = tk.BooleanVar(value=True)
            mode = tk.StringVar(value="To")
            ttk.Checkbutton(single_recips_frame, text=f"{role}: {email}", variable=sel).pack(anchor="w", pady=1)
            ttk.Combobox(single_recips_frame, textvariable=mode, values=["To", "CC", "BCC"], state="readonly", width=6).pack(anchor="w", padx=4, pady=(0, 4))
            canonical = "RegionalEngineeringSpecialist" if role.startswith("Regional") else role
            single_recipient_controls.append((sel, mode, email, canonical))
        else:
            ttk.Label(single_recips_frame, text=f"{role}: N/A", foreground="gray").pack(anchor="w")


def send_single_inline():
    """Draft single email using inline compose UI."""
    if detail_row_current is None:
        messagebox.showinfo("Keine Auswahl", "Bitte zuerst ein Hotel auswaehlen.")
        return
    if os.name != "nt":
        messagebox.showerror("Unsupported Platform", "Outlook email drafting is only available on Windows.")
        return
    if not WIN32COM_AVAILABLE:
        messagebox.showerror("Outlook Not Available", "This feature requires Microsoft Outlook and 'pywin32' (win32com.client).")
        return
    try:
        outlook = get_outlook_app()
        mail_item = outlook.CreateItem(0)
        try:
            mail_item.BodyFormat = 2  # olFormatHTML
        except Exception:
            pass
    except Exception as exc:
        messagebox.showerror("Email Error", f"Could not draft email in Outlook: {exc}")
        return

    to_list, cc_list, bcc_list = [], [], []
    for sel_var, mode_var, email, role_key in single_recipient_controls:
        if sel_var.get() and email:
            for em in normalize_emails(email):
                mode = mode_var.get()
                if mode == "To":
                    to_list.append(em)
                elif mode == "CC":
                    cc_list.append(em)
                elif mode == "BCC":
                    bcc_list.append(em)
    if not (to_list or cc_list or bcc_list):
        messagebox.showinfo("No Recipients", "No recipients selected.")
        return

    mail_item.To = ";".join(to_list)
    mail_item.CC = ";".join(cc_list)
    mail_item.BCC = ";".join(bcc_list)

    subject_template = single_subj_var.get()
    body_template = single_body_text.get("1.0", "end").rstrip("\n")
    mail_item.Subject = render_template(detail_row_current, subject_template)
    sigs = load_signatures()
    sig_entry = sigs.get(single_sig_var.get(), {"html": "", "text": ""})
    rendered = render_with_signature(
        render_template(detail_row_current, body_template),
        sig_entry,
        False,
    )
    if rendered.get("html"):
        mail_item.HTMLBody = rendered["html"]
    else:
        mail_item.Body = rendered.get("text", "")

    attach_enabled = single_attachments_enabled_var.get() if single_attachments_enabled_var else False
    attach_root = single_attachments_root_var.get() if single_attachments_root_var else ""
    if attach_enabled:
        attach_files_for_hotel(mail_item, attach_root, str(detail_row_current.get("Spirit Code", "")).strip())
    mail_item.Display()


def init_single_compose_ui(parent):
    """Build inline single-email compose UI on the lookup tab."""
    global single_compose_frame, single_subj_var, single_body_text, single_sig_var, single_recips_frame
    single_compose_frame = ttk.LabelFrame(parent, text="Create Outlook Draft (Single)", padding=8)
    single_compose_frame.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=(8, 0))

    ttk.Label(single_compose_frame, text="Subject (supports placeholders):").pack(anchor="w", padx=6, pady=(4, 2))
    single_subj_var = tk.StringVar()
    subj_entry = ttk.Entry(single_compose_frame, textvariable=single_subj_var)
    subj_entry.pack(fill="x", padx=6)

    ttk.Label(single_compose_frame, text="Body (supports placeholders):").pack(anchor="w", padx=6, pady=(8, 2))
    single_body_text = tk.Text(single_compose_frame, height=8, font=("Aptos", 12))
    single_body_text.pack(fill="both", expand=True, padx=6)

    link_frame = ttk.Frame(single_compose_frame)
    link_frame.pack(fill="x", padx=6, pady=(2, 4))
    ttk.Button(link_frame, text="Insert Link...", command=lambda: open_link_dialog(single_body_text)).pack(side="left")

    ph_frame = ttk.Frame(single_compose_frame)
    ph_frame.pack(fill="x", padx=6, pady=4)
    ttk.Label(ph_frame, text="Placeholders:").pack(side="left")
    ph_var = tk.StringVar(value=PLACEHOLDERS[0])
    ph_combo = ttk.Combobox(ph_frame, textvariable=ph_var, values=PLACEHOLDERS, state="readonly", width=20)
    ph_combo.pack(side="left", padx=4)

    def insert_placeholder(target="body"):
        ph = ph_var.get()
        if target == "body":
            single_body_text.insert("insert", ph)
        else:
            subj_entry.insert("insert", ph)

    ttk.Button(ph_frame, text="Insert in Body", command=lambda: insert_placeholder("body")).pack(side="left", padx=4)
    ttk.Button(ph_frame, text="Insert in Subject", command=lambda: insert_placeholder("subj")).pack(side="left", padx=4)

    sigs = load_signatures()
    ttk.Label(single_compose_frame, text="Signature:").pack(anchor="w", padx=6, pady=(4, 2))
    single_sig_var = tk.StringVar(value="None")
    ttk.Combobox(single_compose_frame, textvariable=single_sig_var, values=list(sigs.keys()), state="readonly").pack(
        fill="x", padx=6, pady=(0, 6)
    )

    single_recips_frame = ttk.LabelFrame(single_compose_frame, text="Recipients (per email)", padding=6)
    single_recips_frame.pack(fill="x", padx=6, pady=(0, 6))

    ttk.Button(single_compose_frame, text="Create Draft", command=send_single_inline).pack(anchor="e", padx=6, pady=6)

def open_link_dialog(target_text: tk.Text):
    """Open a small dialog to insert a friendly link into the given text widget."""
    dlg = tk.Toplevel(root)
    dlg.title("Insert Link")
    dlg.geometry("360x170")
    dlg.grab_set()

    label_var = tk.StringVar()
    url_var = tk.StringVar(value="https://")

    ttk.Label(dlg, text="Link label:").pack(anchor="w", padx=10, pady=(10, 2))
    ttk.Entry(dlg, textvariable=label_var).pack(fill="x", padx=10)
    ttk.Label(dlg, text="Link URL:").pack(anchor="w", padx=10, pady=(8, 2))
    ttk.Entry(dlg, textvariable=url_var).pack(fill="x", padx=10)

    def insert_link():
        label = label_var.get().strip()
        url = url_var.get().strip()
        if not url:
            messagebox.showinfo("Insert Link", "Bitte geben Sie eine URL ein.")
            return
        if not label:
            label = url
        if not url.lower().startswith(("http://", "https://")):
            url_full = "https://" + url
        else:
            url_full = url
        snippet = f'<a href="{html.escape(url_full)}">{html.escape(label)}</a>'
        target_text.insert("insert", snippet)
        dlg.destroy()

    btn_frame = ttk.Frame(dlg)
    btn_frame.pack(fill="x", padx=10, pady=10)
    ttk.Button(btn_frame, text="Insert", command=insert_link).pack(side="left", padx=4)
    ttk.Button(btn_frame, text="Cancel", command=dlg.destroy).pack(side="right", padx=4)


def ensure_style():
    """Configure ttk style accents (e.g., active tab highlighting)."""
    global style
    if style is None:
        style = ttk.Style()
    current_theme = style.theme_use()
    style.theme_use(current_theme)
    # Blue accent for active tab
    style.map("TNotebook.Tab", background=[("selected", "#1f4fa3")], foreground=[("selected", "white")])
    style.configure("TNotebook.Tab", padding=(8, 4))


def show_splash():
    """Show a splash window while loading."""
    global splash_win, splash_status_var, splash_file_var, splash_logo_img
    if splash_win is not None:
        try:
            splash_win.destroy()
        except Exception:
            pass
    splash_win = tk.Toplevel()
    splash_win.overrideredirect(True)
    splash_win.attributes("-topmost", True)
    splash_win.transient(root)

    container = ttk.Frame(splash_win, padding=14, relief="raised", borderwidth=3)
    container.pack(fill="both", expand=True)

    logo_found = False
    logo_path_use = LOGO_PATH if os.path.isfile(LOGO_PATH) else os.path.join(DATA_DIR, "hyatt_logo.png")
    if os.path.isfile(logo_path_use):
        try:
            img_raw = tk.PhotoImage(file=logo_path_use)
            max_width = 400  # keep logo size reasonable
            if img_raw.width() > max_width:
                factor = math.ceil(img_raw.width() / max_width)
                splash_logo_img = img_raw.subsample(factor, factor)
            else:
                splash_logo_img = img_raw
            ttk.Label(container, image=splash_logo_img).pack(anchor="w", pady=(0, 6))
            logo_found = True
        except Exception:
            logo_found = False
    if not logo_found:
        ttk.Label(container, text="HYATT", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(0, 6))

    ttk.Label(container, text=TOOL_NAME, font=("Segoe UI", 14, "bold")).pack(anchor="w", pady=(10, 3))
    ttk.Label(container, text=f"Version {VERSION} ({VERSION_DATE})", font=("Segoe UI", 11)).pack(anchor="w")
    ttk.Label(container, text="Author: Torsten Heyorth, Dir Engineering Operations", font=("Segoe UI", 10)).pack(anchor="w")
    ttk.Label(container, text="Created with OpenAI Codex & VS Code", font=("Segoe UI", 10)).pack(anchor="w", pady=(0, 10))

    splash_file_var = tk.StringVar(value="Loading data file...")
    ttk.Label(container, textvariable=splash_file_var, font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(4, 2))

    splash_status_var = tk.StringVar(value=f"{TOOL_NAME} v{VERSION} ({VERSION_DATE})")
    ttk.Label(container, textvariable=splash_status_var, foreground="gray", font=("Segoe UI", 10)).pack(anchor="w")

    ttk.Button(container, text="Understood...", command=close_splash).pack(anchor="e", pady=(12, 0), ipadx=8, ipady=3)

    splash_win.update_idletasks()
    desired_w = 720
    desired_h = 420
    ws = splash_win.winfo_screenwidth()
    hs = splash_win.winfo_screenheight()
    x = int((ws / 2) - (desired_w / 2))
    y = int((hs / 2) - (desired_h / 2))
    splash_win.geometry(f"{desired_w}x{desired_h}+{x}+{y}")
    splash_win.config(highlightbackground="#1f4fa3", highlightcolor="#1f4fa3", highlightthickness=2)
    try:
        splash_win.lift()
        splash_win.focus_force()
        splash_win.after(50, splash_win.lift)
        splash_win.after(50, lambda: splash_win.geometry(f"{desired_w}x{desired_h}+{x}+{y}"))
    except Exception:
        pass


def update_splash(file_path: str, status: str):
    if splash_file_var is not None:
        if file_path:
            splash_file_var.set(f"Loaded file: {os.path.basename(file_path)}")
        else:
            splash_file_var.set("Loading data file...")
    if splash_status_var is not None and status:
        splash_status_var.set(status)
    if splash_win is not None:
        splash_win.lift()
        splash_win.after(50, splash_win.lift)


def close_splash():
    """Close splash window and then prompt for config once."""
    global splash_win
    if splash_win is not None:
        try:
            splash_win.destroy()
        except Exception:
            pass
    splash_win = None
    maybe_prompt_config_after_splash()


def maybe_prompt_config_after_splash():
    """Show startup config prompt only once after splash is closed."""
    global config_prompted
    if config_prompted:
        return
    config_prompted = True
    prompt_startup_config()


def refresh_filter_columns_list():
    """Refresh list of selectable columns for the filtered hotels table."""
    if filter_cols_listbox is None:
        return
    filter_cols_listbox.delete(0, tk.END)
    candidates = []
    if not df.empty:
        candidates = [c for c in sorted(df.columns) if c not in MANDATORY_FILTER_COLS]
    else:
        candidates = [c for c in visible_optional_filter_cols if c not in MANDATORY_FILTER_COLS]
    for col in candidates:
        filter_cols_listbox.insert(tk.END, col)
    # restore selections
    for idx, col in enumerate(candidates):
        if col in visible_optional_filter_cols:
            filter_cols_listbox.selection_set(idx)


def get_filtered_columns():
    """Return ordered list of columns to show in filtered tree (mandatory + selected optional)."""
    selected = []
    if filter_cols_listbox is not None:
        selected = [filter_cols_listbox.get(i) for i in filter_cols_listbox.curselection()]
    if not selected:
        selected = visible_optional_filter_cols
    return MANDATORY_FILTER_COLS + selected


def add_role_selector(parent, role_name, default_mode="Skip"):
    var = tk.StringVar(value=default_mode)
    cb = ttk.Combobox(parent, textvariable=var, values=ROLE_MODES, state="readonly", width=10)
    cb.bind("<<ComboboxSelected>>", lambda e: update_selected_tree())
    role_send_vars[role_name] = var
    row = len(parent.grid_slaves()) // 2
    ttk.Label(parent, text=role_name).grid(row=row, column=0, sticky="w", padx=4, pady=2)
    cb.grid(row=row, column=1, sticky="w", padx=4, pady=2)


def get_country_value(row: pd.Series) -> str:
    """Return the country/area value from the configured columns."""
    country_col = get_country_col()
    fallback_col = get_country_fallback_col()
    if country_col and country_col in row and pd.notna(row[country_col]):
        return str(row[country_col])
    if fallback_col and fallback_col in row and pd.notna(row[fallback_col]):
        return str(row[fallback_col])
    return ""


def get_city_value(row: pd.Series) -> str:
    city_col = get_city_col()
    if city_col and city_col in row and pd.notna(row[city_col]):
        return str(row[city_col])
    return ""


def update_status():
    """Refresh status line and file label with current metadata."""
    global status_var
    if status_var is None:
        return

    if data_file_path and os.path.isfile(data_file_path):
        hotel_count = len(df) if not df.empty else 0
        status_var.set(
            f"Datei: {os.path.basename(data_file_path)} | Stand: {format_timestamp(data_file_path)} | Hotels geladen: {hotel_count}"
        )
    else:
        status_var.set("Keine Datendatei geladen")


def ensure_var_in_columns(var: tk.StringVar, preferred_order: list[str], allow_none: bool = False):
    """Ensure a column selection variable is set to an available column or None."""
    if var is None:
        return
    cols = list(df.columns)
    current = var.get()
    if current in cols or (allow_none and current == "None"):
        return
    for candidate in preferred_order:
        if candidate and candidate in cols:
            var.set(candidate)
            return
    if allow_none:
        var.set("None")
    elif cols:
        var.set(cols[0])
    else:
        var.set("")


def refresh_setup_tab_options():
    """Populate setup tab combos with current dataframe columns and keep selections valid."""
    cols = sorted([c for c in df.columns]) if not df.empty else []
    default_list = cols or [""]

    for combo in [
        brand_col_combo,
        region_col_combo,
        city_col_combo,
        gm_col_combo,
        eng_col_combo,
        dof_col_combo,
        avp_col_combo,
        md_col_combo,
        brand_band_col_combo,
        hyatt_date_col_combo,
        relationship_col_combo,
    ]:
        if combo is not None:
            combo["values"] = ["None"] + default_list

    if country_col_combo is not None:
        country_col_combo["values"] = ["None"] + default_list
    if country_fallback_combo is not None:
        country_fallback_combo["values"] = ["None"] + default_list
    if reg_eng_spec_combo is not None:
        reg_eng_spec_combo["values"] = ["None"] + default_list

    ensure_var_in_columns(brand_col_var, [DEFAULT_BRAND_COL], allow_none=True)
    ensure_var_in_columns(region_col_var, [DEFAULT_REGION_COL], allow_none=True)
    ensure_var_in_columns(city_col_var, [DEFAULT_CITY_COL], allow_none=True)
    ensure_var_in_columns(brand_band_col_var, [DEFAULT_BRAND_BAND_COL], allow_none=True)
    ensure_var_in_columns(relationship_col_var, [DEFAULT_RELATIONSHIP_COL], allow_none=True)
    ensure_var_in_columns(country_col_var, [DEFAULT_COUNTRY_COL, DEFAULT_COUNTRY_FALLBACK_COL], allow_none=True)
    ensure_var_in_columns(country_fallback_col_var, [DEFAULT_COUNTRY_FALLBACK_COL], allow_none=True)
    ensure_var_in_columns(hyatt_date_col_var, [DEFAULT_HYATT_DATE_COL], allow_none=True)
    ensure_var_in_columns(gm_col_var, [DEFAULT_GM_COL, "GM"], allow_none=True)  # include old fallback GM
    ensure_var_in_columns(eng_col_var, [DEFAULT_ENG_COL, "Engineering Director"], allow_none=True)  # include old fallback Engineering Director
    ensure_var_in_columns(dof_col_var, [DEFAULT_DOF_COL], allow_none=True)
    ensure_var_in_columns(avp_col_var, ["AVP of Ops", "AVP of Ops-managed"], allow_none=True)
    ensure_var_in_columns(md_col_var, ["SVP / Managing Director", "SVP"], allow_none=True)
    ensure_var_in_columns(reg_eng_spec_col_var, [DEFAULT_REG_ENG_SPEC_COL], allow_none=True)


def apply_column_settings():
    """Apply column selections to filters and refresh views."""
    update_visible_optional_from_listbox()
    update_filter_options()
    refresh_filtered_hotels()
    update_selected_tree()


def update_filter_options():
    """Populate filter dropdowns based on loaded data and chosen columns."""
    global brand_values, region_values, country_values, brand_band_values, relationship_values
    brand_col = get_brand_col()
    region_col = get_region_col()
    country_col = get_country_col() or get_country_fallback_col()
    brand_band_col = get_brand_band_col()
    relationship_col = get_relationship_col()

    if df.empty:
        brand_values = []
        region_values = []
        country_values = []
        brand_band_values = []
        relationship_values = []
    else:
        brand_values = sorted(df[brand_col].dropna().astype(str).unique().tolist()) if brand_col in df.columns else []
        region_values = sorted(df[region_col].dropna().astype(str).unique().tolist()) if region_col in df.columns else []
        if country_col and country_col in df.columns:
            country_values = sorted(df[country_col].dropna().astype(str).unique().tolist())
        else:
            country_values = []
        brand_band_values = (
            sorted(df[brand_band_col].dropna().astype(str).unique().tolist()) if brand_band_col in df.columns else []
        )
        relationship_values = (
            sorted(df[relationship_col].dropna().astype(str).unique().tolist()) if relationship_col in df.columns else []
        )

    if brand_filter_var is not None:
        brand_filter_var.set("Any")
    if region_filter_var is not None:
        region_filter_var.set("Any")
    if country_filter_var is not None:
        country_filter_var.set("Any")
    if brand_band_filter_var is not None:
        brand_band_filter_var.set("Any")
    if relationship_filter_var is not None:
        relationship_filter_var.set("Any")

    def reset_listbox(lb, values):
        if lb is None:
            return
        lb.delete(0, tk.END)
        for val in values:
            lb.insert(tk.END, val)

    reset_listbox(brand_listbox, brand_values)
    reset_listbox(brand_band_listbox, brand_band_values)
    reset_listbox(region_listbox, region_values)
    reset_listbox(relationship_listbox, relationship_values)
    reset_listbox(country_listbox, country_values)

    refresh_filter_columns_list()

    if filtered_tree is not None:
        refresh_filtered_hotels()


def update_visible_optional_from_listbox():
    """Capture selected optional columns for the filtered hotels view."""
    global visible_optional_filter_cols
    if filter_cols_listbox is None:
        return
    selected = [filter_cols_listbox.get(i) for i in filter_cols_listbox.curselection()]
    if selected:
        visible_optional_filter_cols = selected


def reset_filters():
    """Clear all filter selections."""
    for lb in [brand_listbox, brand_band_listbox, region_listbox, relationship_listbox, country_listbox]:
        if lb is not None:
            lb.selection_clear(0, tk.END)
    if hyatt_year_var is not None:
        hyatt_year_var.set("")
    if hyatt_year_mode_var is not None:
        hyatt_year_mode_var.set("Any")
    refresh_filtered_hotels()


def load_data(path: str):
    """Load Excel data and refresh UI widgets."""
    global df, hotel_names, data_file_path

    new_df = pd.read_excel(path, engine="openpyxl")
    if "Hotel" not in new_df.columns:
        raise ValueError("Die ausgewaehlte Datei enthaelt keine Spalte 'Hotel'.")

    df = new_df
    hotel_names = sorted(df["Hotel"].dropna().unique().tolist())
    data_file_path = path

    if hotel_combo is not None:
        hotel_combo["values"] = hotel_names

    refresh_setup_tab_options()
    update_filter_options()
    update_status()
    update_splash(path, "Data loaded.")


def prompt_for_file():
    """Ask user to select an Excel file and load it."""
    initial_dir = DATA_DIR if os.path.isdir(DATA_DIR) else os.getcwd()
    file_path = filedialog.askopenfilename(
        title="Excel-Datei auswaehlen",
        initialdir=initial_dir,
        filetypes=[("Excel-Dateien", "*.xlsx *.xlsm *.xls"), ("Alle Dateien", "*.*")],
    )
    if not file_path:
        return
    update_splash("", f"Loading {os.path.basename(file_path)} ...")
    try:
        load_data(file_path)
        update_splash(file_path, "Data loaded.")
    except Exception as exc:
        messagebox.showerror("Laden fehlgeschlagen", f"Die Datei konnte nicht geladen werden:\n{exc}")


def load_recent_configs():
    """Return list of recent configuration file paths."""
    if not os.path.isfile(RECENT_CONFIG_PATH):
        return []
    try:
        with open(RECENT_CONFIG_PATH, "r", encoding="utf-8") as fh:
            paths = json.load(fh)
        return [p for p in paths if isinstance(p, str) and p]
    except Exception:
        return []


def save_recent_configs(paths: list[str]):
    """Persist recent configuration file paths (max 8)."""
    uniq = []
    seen = set()
    for p in paths:
        if not p or not isinstance(p, str):
            continue
        if p in seen:
            continue
        uniq.append(p)
        seen.add(p)
    uniq = uniq[:8]
    try:
        with open(RECENT_CONFIG_PATH, "w", encoding="utf-8") as fh:
            json.dump(uniq, fh, indent=2)
    except Exception:
        pass


def remember_config(path: str):
    """Store path in recent config list."""
    recents = load_recent_configs()
    recents = [p for p in recents if p != path]
    recents.insert(0, path)
    save_recent_configs(recents)


def load_config_from_path(config_path: str):
    """Load configuration JSON from explicit path (no dialogs)."""
    if not config_path:
        return
    try:
        with open(config_path, "r", encoding="utf-8") as fh:
            cfg = json.load(fh)
    except Exception as exc:
        messagebox.showerror("Konfiguration", f"Konfigurationsdatei konnte nicht gelesen werden:\n{exc}")
        return

    data_path = cfg.get("data_file_path")
    if data_path:
        try:
            load_data(data_path)
        except Exception as exc:
            messagebox.showerror("Konfiguration", f"Datendatei aus Konfiguration konnte nicht geladen werden:\n{exc}")

    cols = cfg.get("columns", {})

    def set_if_present(var, key):
        if var is not None and key in cols and cols.get(key):
            var.set(cols[key])

    set_if_present(brand_col_var, "brand")
    set_if_present(brand_band_col_var, "brand_band")
    set_if_present(region_col_var, "region")
    set_if_present(relationship_col_var, "relationship")
    set_if_present(country_col_var, "country")
    set_if_present(country_fallback_col_var, "country_fallback")
    set_if_present(city_col_var, "city")
    set_if_present(hyatt_date_col_var, "hyatt_date")
    set_if_present(gm_col_var, "gm")
    set_if_present(eng_col_var, "eng")
    set_if_present(dof_col_var, "dof")
    set_if_present(avp_col_var, "avp")
    set_if_present(md_col_var, "md")
    set_if_present(reg_eng_spec_col_var, "reg_eng_spec")

    roles_cfg = cfg.get("roles", {})
    for role, val in roles_cfg.items():
        if role in role_send_vars and val in ROLE_MODES:
            role_send_vars[role].set(val)

    optional_cols = cfg.get("visible_filter_cols")
    global visible_optional_filter_cols
    if optional_cols and isinstance(optional_cols, list):
        visible_optional_filter_cols = optional_cols
        refresh_filter_columns_list()
        refresh_filtered_hotels()

    attach_cfg = cfg.get("attachments", {})
    if attachments_enabled_var is not None and "enabled" in attach_cfg:
        attachments_enabled_var.set(bool(attach_cfg.get("enabled", False)))
    if attachments_root_var is not None and "root" in attach_cfg:
        attachments_root_var.set(attach_cfg.get("root", ""))

    refresh_setup_tab_options()
    apply_column_settings()
    update_selected_tree()

    remember_config(config_path)
    update_status()
    update_splash(config_path, f"Konfiguration geladen: {os.path.basename(config_path)}")


def load_config_file():
    """Load configuration JSON (data file path and role routing)."""
    config_path = filedialog.askopenfilename(
        title="Konfiguration laden",
        filetypes=[("JSON files", "*.json"), ("Alle Dateien", "*.*")],
    )
    if not config_path:
        return
    load_config_from_path(config_path)


def save_config_file():
    """Save configuration (data file path, column mapping, role routing) to JSON."""
    update_visible_optional_from_listbox()
    config_path = filedialog.asksaveasfilename(
        title="Konfiguration speichern",
        defaultextension=".json",
        filetypes=[("JSON files", "*.json"), ("Alle Dateien", "*.*")],
    )
    if not config_path:
        return

    cfg = {
        "data_file_path": data_file_path,
        "columns": {
            "brand": get_brand_col(),
            "brand_band": get_brand_band_col(),
            "region": get_region_col(),
            "relationship": get_relationship_col(),
            "country": get_country_col(),
            "country_fallback": get_country_fallback_col(),
            "city": get_city_col(),
            "hyatt_date": get_hyatt_date_col(),
            "gm": get_gm_col(),
            "eng": get_eng_col(),
            "dof": get_dof_col(),
            "avp": get_avp_col(),
            "md": get_md_col(),
            "reg_eng_spec": get_reg_eng_spec_col(),
        },
        "roles": {role: var.get() for role, var in role_send_vars.items()},
        "visible_filter_cols": visible_optional_filter_cols,
        "attachments": {
            "enabled": attachments_enabled_var.get() if attachments_enabled_var else False,
            "root": attachments_root_var.get() if attachments_root_var else "",
            "common_dir": attachments_common_dir,
            "spirit_dir": attachments_spirit_dir,
        },
    }
    try:
        with open(config_path, "w", encoding="utf-8") as fh:
            json.dump(cfg, fh, indent=2)
    except Exception as exc:
        messagebox.showerror("Konfiguration", f"Konfiguration konnte nicht gespeichert werden:\n{exc}")


def ensure_initial_data():
    """Initialize without auto-loading data; user will choose a configuration at startup."""
    update_splash("", f"{TOOL_NAME} v{VERSION} ({VERSION_DATE}) - bitte Konfiguration laden...")
    update_status()


def prompt_startup_config():
    """Prompt the user to load a configuration on startup (optional)."""
    dlg = tk.Toplevel(root)
    dlg.title("Konfiguration laden")
    dlg.geometry("520x220")
    dlg.grab_set()

    ttk.Label(dlg, text="Konfiguration laden (optional):").pack(anchor="w", padx=10, pady=(10, 4))
    path_var = tk.StringVar()
    recent_configs = load_recent_configs()
    if recent_configs:
        path_var.set(recent_configs[0])
    ttk.Entry(dlg, textvariable=path_var).pack(fill="x", padx=10, pady=(0, 6))
    if recent_configs:
        ttk.Label(dlg, text="Zuletzt verwendet:").pack(anchor="w", padx=10)
        ttk.Combobox(dlg, values=recent_configs, textvariable=path_var, state="readonly").pack(fill="x", padx=10, pady=(0, 6))

    def browse_config():
        cfg = filedialog.askopenfilename(
            title="Konfiguration laden",
            filetypes=[("JSON files", "*.json"), ("Alle Dateien", "*.*")],
        )
        if cfg:
            path_var.set(cfg)

    def load_selected():
        path = path_var.get().strip()
        if not path:
            messagebox.showinfo("Konfiguration", "Bitte waehlen Sie eine Konfigurationsdatei oder fahren Sie ohne fort.")
            return
        load_config_from_path(path)
        close_splash()
        dlg.destroy()

    def skip():
        close_splash()
        dlg.destroy()

    btn_frame = ttk.Frame(dlg)
    btn_frame.pack(fill="x", padx=10, pady=10)
    ttk.Button(btn_frame, text="Durchsuchen...", command=browse_config).pack(side="left", padx=4)
    ttk.Button(btn_frame, text="Laden", command=load_selected).pack(side="left", padx=4)
    ttk.Button(btn_frame, text="Ohne Konfiguration fortfahren", command=skip).pack(side="right", padx=4)

def get_outlook_app(force_refresh: bool = False):
    """Return a cached Outlook Application COM object (or create it on first use)."""
    global _outlook_app

    if _outlook_app is not None and not force_refresh:
        return _outlook_app

    import win32com.client  # type: ignore[import-untyped]

    try:
        _outlook_app = win32com.client.gencache.EnsureDispatch("Outlook.Application")
    except Exception:
        # Handle broken gencache (e.g., CLSIDToPackageMap issues)
        try:
            gen_path = win32com.client.gencache.GetGeneratePath()
            if os.path.isdir(gen_path):
                shutil.rmtree(gen_path, ignore_errors=True)
        except Exception:
            pass
        _outlook_app = win32com.client.Dispatch("Outlook.Application")

    return _outlook_app


def warm_outlook_app():
    """Preload Outlook once to speed up the first email draft."""
    if os.name == "nt" and WIN32COM_AVAILABLE:
        try:
            get_outlook_app()
            update_splash(data_file_path, "Warming Outlook...")
        except Exception:
            pass


# ---------------------------------------------------------------------------
# Lookup detail panel helpers
# ---------------------------------------------------------------------------
def init_detail_panel(parent):
    """Build the detail panel on the lookup tab."""
    global detail_info_vars, detail_roles_frame, detail_status_var, detail_start_email_btn

    detail_frame = ttk.LabelFrame(parent, text="Hotel Details", padding=10)
    detail_frame.pack(fill="both", expand=True)

    info_grid = ttk.Frame(detail_frame)
    info_grid.pack(fill="x", pady=(0, 8))

    fields = [
        "Spirit Code",
        "Hotel",
        "City",
        "Relationship",
        "Brand",
        "Brand Band",
        "Region",
        "Country/Area",
    ]
    detail_info_vars = {name: tk.StringVar(value="") for name in fields}

    for idx, name in enumerate(fields):
        ttk.Label(info_grid, text=f"{name}:").grid(row=idx, column=0, sticky="w", padx=4, pady=2)
        ttk.Label(info_grid, textvariable=detail_info_vars[name], width=35).grid(
            row=idx, column=1, sticky="w", padx=4, pady=2
        )

    detail_roles_frame = ttk.LabelFrame(detail_frame, text="Recipients", padding=8)
    detail_roles_frame.pack(fill="both", expand=True, pady=(0, 8))

    ttk.Label(detail_roles_frame, text="Recipients are configured in the composer below.", foreground="gray").pack(
        anchor="w"
    )

    detail_status_var = tk.StringVar(value="Select a hotel to view details.")
    ttk.Label(detail_frame, textvariable=detail_status_var, foreground="gray").pack(anchor="w", pady=(4, 0))


def clear_detail_panel(message: str = "Select a hotel to view details."):
    """Reset detail panel contents."""
    global detail_checkbox_vars, detail_hotel_name
    detail_checkbox_vars = []
    detail_hotel_name = ""
    for var in detail_info_vars.values():
        var.set("")
    for widget in detail_roles_frame.winfo_children():
        widget.destroy()
    ttk.Label(detail_roles_frame, text="Recipients are configured in the composer below.", foreground="gray").pack(anchor="w")
    if detail_status_var is not None:
        detail_status_var.set(message)
    clear_single_compose()
    clear_single_compose()


def populate_detail_panel(row: pd.Series):
    """Fill detail panel with hotel info and role checkboxes."""
    global detail_checkbox_vars, detail_hotel_name, detail_row_current
    detail_checkbox_vars = []
    detail_hotel_name = row.get("Hotel", "N/A")
    detail_row_current = row

    if detail_status_var is not None:
        detail_status_var.set(f"Details loaded for: {detail_hotel_name}")

    city_val = get_city_value(row)
    relationship_col = get_relationship_col()
    brand_col = get_brand_col()
    brand_band_col = get_brand_band_col()
    region_col = get_region_col()

    detail_info_vars["Spirit Code"].set(row.get("Spirit Code", ""))
    detail_info_vars["Hotel"].set(detail_hotel_name)
    detail_info_vars["City"].set(city_val)
    detail_info_vars["Relationship"].set(row.get(relationship_col, "") if relationship_col in row else "")
    detail_info_vars["Brand"].set(row.get(brand_col, "") if brand_col in row else "")
    detail_info_vars["Brand Band"].set(row.get(brand_band_col, "") if brand_band_col in row else "")
    detail_info_vars["Region"].set(row.get(region_col, "") if region_col in row else "")
    detail_info_vars["Country/Area"].set(get_country_value(row))

    for widget in detail_roles_frame.winfo_children():
        widget.destroy()
    ttk.Label(detail_roles_frame, text="Recipients are configured in the composer below.", foreground="gray").pack(anchor="w")
    update_single_compose(row)

# ---------------------------------------------------------------------------
# Multi-select helpers
# ---------------------------------------------------------------------------
def filtered_dataframe():
    """Return dataframe filtered by current dropdown selections."""
    if df.empty:
        return pd.DataFrame()

    filt = df
    brand_col = get_brand_col()
    region_col = get_region_col()
    country_col = get_country_col() or get_country_fallback_col()
    brand_band_col = get_brand_band_col()
    hyatt_col = get_hyatt_date_col()
    relationship_col = get_relationship_col()

    def selected_values(lb):
        if lb is None:
            return []
        return [lb.get(i) for i in lb.curselection()]

    selected_brands = selected_values(brand_listbox)
    selected_regions = selected_values(region_listbox)
    selected_countries = selected_values(country_listbox)
    selected_bands = selected_values(brand_band_listbox)
    selected_relationships = selected_values(relationship_listbox)
    quick_codes = []
    if quick_spirit_var is not None and quick_spirit_var.get().strip():
        quick_codes = [c.strip() for c in quick_spirit_var.get().split(",") if c.strip()]

    if selected_brands and brand_col in filt.columns:
        filt = filt[filt[brand_col].astype(str).isin(selected_brands)]
    if selected_regions and region_col in filt.columns:
        filt = filt[filt[region_col].astype(str).isin(selected_regions)]
    if selected_countries and country_col in filt.columns:
        filt = filt[filt[country_col].astype(str).isin(selected_countries)]
    if selected_bands and brand_band_col in filt.columns:
        filt = filt[filt[brand_band_col].astype(str).isin(selected_bands)]
    if selected_relationships and relationship_col in filt.columns:
        filt = filt[filt[relationship_col].astype(str).isin(selected_relationships)]

    # Hyatt date filter (year with before/after/on)
    if hyatt_col and hyatt_col in filt.columns and hyatt_year_mode_var is not None and hyatt_year_var is not None:
        mode = hyatt_year_mode_var.get()
        year_str = hyatt_year_var.get().strip()
        if mode and mode != "Any" and year_str.isdigit():
            target_year = int(year_str)
            years = pd.to_datetime(filt[hyatt_col], errors="coerce").dt.year
            if mode == "Before":
                filt = filt[years.notna() & (years < target_year)]
            elif mode == "Before/Equal":
                filt = filt[years.notna() & (years <= target_year)]
            elif mode == "Equal":
                filt = filt[years.notna() & (years == target_year)]
            elif mode == "After/Equal":
                filt = filt[years.notna() & (years >= target_year)]
            elif mode == "After":
                filt = filt[years.notna() & (years > target_year)]
    if quick_codes and "Spirit Code" in filt.columns:
        filt = filt[filt["Spirit Code"].astype(str).isin(quick_codes)]
    return filt


def refresh_filtered_hotels():
    """Refresh the filtered hotels list."""
    global current_filtered_indexes
    if filtered_tree is None:
        return

    for item in filtered_tree.get_children():
        filtered_tree.delete(item)

    filt_df = filtered_dataframe()
    hyatt_col_name = get_hyatt_date_col()
    current_filtered_indexes = []

    # Reconfigure columns based on selection
    cols = get_filtered_columns()
    filtered_tree["columns"] = cols
    for col in cols:
        filtered_tree.heading(col, text=col)
        filtered_tree.column(col, width=120, stretch=True)
    enable_treeview_sort(filtered_tree)

    for idx, (_, row) in enumerate(filt_df.iterrows()):
        tree_id = str(row.name)
        current_filtered_indexes.append(row.name)
        values = []
        for col in cols:
            if col == "Spirit Code":
                values.append(row.get("Spirit Code", ""))
            elif col == "Hotel":
                values.append(row.get("Hotel", ""))
            elif col == "Country" or col == "Country/Area":
                values.append(get_country_value(row))
            elif col == "City":
                values.append(get_city_value(row))
            elif hyatt_col_name and col == hyatt_col_name:
                dt_val = pd.to_datetime(row.get(col, ""), errors="coerce")
                if pd.notna(dt_val):
                    values.append(dt_val.strftime("%Y-%m"))
                else:
                    values.append(str(row.get(col, "")))
            else:
                values.append(row.get(col, ""))
        filtered_tree.insert("", "end", iid=tree_id, values=tuple(values))

    if filtered_count_var is not None:
        filtered_count_var.set(f"Filtered: {len(filt_df)}")


def add_selected_hotels():
    """Add selected rows from the filtered tree to the selection list."""
    if filtered_tree is None:
        return
    selected = filtered_tree.selection()
    if not selected:
        messagebox.showinfo("Auswahl", "Bitte waehlen Sie mindestens ein Hotel aus der Filterliste aus.")
        return

    for tree_id in selected:
        try:
            row_idx = int(tree_id)
        except ValueError:
            continue
        if row_idx in selected_rows:
            continue
        row = df.loc[row_idx]
        selected_rows[row_idx] = row
    update_selected_tree()


def remove_selected_hotels():
    """Remove selected rows from the selection list."""
    if selected_tree is None:
        return
    chosen = selected_tree.selection()
    for tree_id in chosen:
        try:
            row_idx = int(tree_id)
        except ValueError:
            continue
        selected_rows.pop(row_idx, None)
    update_selected_tree()

def clear_selected_hotels():
    """Clear all selections."""
    selected_rows.clear()
    update_selected_tree()


def add_all_filtered_hotels():
    """Add all currently filtered hotels to the selection list."""
    filt_df = filtered_dataframe()
    for idx, row in filt_df.iterrows():
        if idx in selected_rows:
            continue
        selected_rows[idx] = row
    update_selected_tree()


def update_selected_tree():
    """Refresh the selected hotels list."""
    if selected_tree is None:
        return
    for item in selected_tree.get_children():
        selected_tree.delete(item)

    for row_idx, row in selected_rows.items():
        availability = {}
        recipients_all = []
        for role, var in role_send_vars.items():
            if var.get() == "Skip":
                availability[role] = "No"
                continue
            addrs = get_role_addresses(row, role)
            if addrs:
                availability[role] = "Yes"
                recipients_all.extend(addrs)
            else:
                availability[role] = "No"
        recipients_all = [r for r in recipients_all if r]

        def mark(role_key: str) -> str:
            return "✓" if availability.get(role_key, "No") == "Yes" else "✗"

        selected_tree.insert(
            "",
            "end",
            iid=str(row_idx),
            values=(
                row.get("Spirit Code", ""),
                row.get("Hotel", ""),
                "; ".join(recipients_all),
                mark("AVP"),
                mark("MD"),
                mark("GM"),
                mark("Engineering"),
                mark("DOF"),
                mark("RegionalEngineeringSpecialist"),
            ),
            tags=(),
        )


def get_role_addresses(row: pd.Series, role_key: str):
    """Return a list of email addresses for the chosen role."""
    role_map = {
        "AVP": [get_avp_col()],
        "MD": [get_md_col()],
        "GM": [get_gm_col()],
        "Engineering": [get_eng_col()],
        "DOF": [get_dof_col()],
        "RegionalEngineeringSpecialist": [get_reg_eng_spec_col()],
    }
    emails = []
    for col in role_map.get(role_key, []):
        if col and col in row and pd.notna(row[col]):
            raw = str(row[col])
            for email in normalize_emails(raw):
                emails.append(email)
    return emails


def bind_autofit(tree: ttk.Treeview, min_width: int = 60):
    """Bind a resize handler to auto-distribute column widths."""
    if tree is None:
        return

    def _on_config(event):
        cols = tree["columns"]
        if not cols:
            return
        total = max(event.width - 20, len(cols) * min_width)
        per = total // len(cols)
        for col in cols:
            tree.column(col, width=per, stretch=True)

    tree.bind("<Configure>", _on_config)


def enable_treeview_sort(tree: ttk.Treeview):
    """Enable clickable column headers for sorting."""
    if tree is None:
        return
    sort_state = {}

    def sort_by(col: str):
        descending = sort_state.get(col, False)
        data = []
        for iid in tree.get_children(""):
            val = tree.set(iid, col)
            def norm(v):
                if v is None:
                    return (2, "")
                s = str(v)
                try:
                    return (0, float(s))
                except ValueError:
                    return (1, s.lower())
            data.append((norm(val), iid))
        data.sort(key=lambda t: t[0], reverse=descending)
        for idx, (_, iid) in enumerate(data):
            tree.move(iid, "", idx)
        sort_state[col] = not descending

    for col in tree["columns"]:
        tree.heading(col, command=lambda c=col: sort_by(c))


def html_table_from_excel_row(row: pd.Series, body_cols: list[str]) -> str:
    """Build a two-column HTML table (header/value) from selected Excel columns."""
    if not body_cols:
        return ""
    rows_html = ""
    for col in body_cols:
        val = row.get(col, "")
        rows_html += f"<tr><td style='padding:4px 8px; font-weight:bold;'>{html.escape(str(col))}</td><td style='padding:4px 8px;'>{html.escape(str(val))}</td></tr>"
    return f"<table border='1' cellspacing='0' cellpadding='0' style='border-collapse:collapse; margin-top:8px;'>{rows_html}</table>"


def prompt_excel_sheet_table(path: str):
    """Prompt user to choose sheet and optionally an Excel table."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        messagebox.showerror("Excel laden", f"Workbook konnte nicht geoeffnet werden:\n{exc}")
        return None, None
    sheets = wb.sheetnames
    tables_by_sheet = {}
    for sname in sheets:
        try:
            ws = wb[sname]
            tables_by_sheet[sname] = list(ws.tables.keys())
        except Exception:
            tables_by_sheet[sname] = []

    dlg = tk.Toplevel(root)
    dlg.title("Excel Sheet/Tabellen Auswahl")
    dlg.geometry("420x180")
    dlg.grab_set()

    ttk.Label(dlg, text="Sheet auswaehlen:").pack(anchor="w", padx=8, pady=(8, 2))
    sheet_var = tk.StringVar(value=sheets[0] if sheets else "")
    sheet_combo = ttk.Combobox(dlg, values=sheets, textvariable=sheet_var, state="readonly")
    sheet_combo.pack(fill="x", padx=8)

    ttk.Label(dlg, text="Excel-Tabelle (optional):").pack(anchor="w", padx=8, pady=(8, 2))
    table_var = tk.StringVar(value="None")
    table_combo = ttk.Combobox(dlg, values=["None"], textvariable=table_var, state="readonly")
    table_combo.pack(fill="x", padx=8)

    def refresh_tables(*args):
        s = sheet_var.get()
        items = ["None"] + tables_by_sheet.get(s, [])
        table_combo["values"] = items
        if table_var.get() not in items:
            table_var.set("None")

    sheet_combo.bind("<<ComboboxSelected>>", refresh_tables)
    refresh_tables()

    choice = {"sheet": None, "table": None}

    def confirm():
        choice["sheet"] = sheet_var.get().strip()
        tbl = table_var.get().strip()
        if tbl == "None":
            tbl = None
        choice["table"] = tbl
        dlg.destroy()

    def cancel():
        choice["sheet"] = None
        choice["table"] = None
        dlg.destroy()

    btns = ttk.Frame(dlg)
    btns.pack(fill="x", padx=8, pady=10)
    ttk.Button(btns, text="Abbrechen", command=cancel).pack(side="right", padx=4)
    ttk.Button(btns, text="OK", command=confirm).pack(side="right", padx=4)

    dlg.wait_window()
    return choice["sheet"], choice["table"]


def load_excel_email_file():
    """Load an Excel list for email generation."""
    global excel_df, excel_mapping, excel_selected_cache
    path = filedialog.askopenfilename(
        title="Excel-Liste laden",
        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
    )
    if not path:
        return
    sheet_name, table_name = prompt_excel_sheet_table(path)
    if not sheet_name:
        return
    try:
        if table_name:
            # Read only the table range
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            ws = wb[sheet_name]
            tbl = ws.tables[table_name]
            min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)
            usecols = f"{get_column_letter(min_col)}:{get_column_letter(max_col)}"
            nrows = max_row - min_row + 1
            excel_df = pd.read_excel(
                path,
                sheet_name=sheet_name,
                engine="openpyxl",
                header=0,
                skiprows=min_row - 1,
                nrows=nrows,
                usecols=usecols,
            )
        else:
            excel_df = pd.read_excel(path, engine="openpyxl", sheet_name=sheet_name)
    except Exception as exc:
        messagebox.showerror("Excel laden", f"Datei konnte nicht geladen werden:\n{exc}")
        return
    if excel_file_label_var is not None:
        label = os.path.basename(path)
        if table_name:
            label += f" [{sheet_name}/{table_name}]"
        else:
            label += f" [{sheet_name}]"
        excel_file_label_var.set(label)
    if excel_df.empty:
        messagebox.showinfo("Excel laden", "Die Excel-Datei ist leer.")
        return
    # Initialize mapping defaults
    new_map = {}
    for col in excel_df.columns:
        new_map[col] = excel_mapping.get(col, "Skip")
    excel_mapping = new_map
    excel_selected_cache = {}
    render_excel_mapping_ui()
    refresh_excel_filtered_tree()


def render_excel_mapping_ui():
    """Show header list with dropdowns for role selection."""
    global excel_mapping_controls, excel_filter_controls
    if excel_headers_frame is None:
        return
    for child in excel_headers_frame.winfo_children():
        child.destroy()
    excel_mapping_controls = []
    excel_filter_controls = []
    options = ["Skip", "Spirit Code", "Include in Body"]
    for idx, col in enumerate(excel_df.columns):
        ttk.Label(excel_headers_frame, text=col).grid(row=idx, column=0, sticky="w", padx=4, pady=2)
        var = tk.StringVar(value=excel_mapping.get(col, "Skip"))
        cb = ttk.Combobox(excel_headers_frame, values=options, textvariable=var, state="readonly", width=18)
        cb.grid(row=idx, column=1, sticky="w", padx=4, pady=2)
        excel_mapping_controls.append((col, var))
        # per-column filter
        ttk.Label(excel_headers_frame, text="Filter").grid(row=idx, column=2, sticky="e", padx=4, pady=2)
        vals = ["<blank>"]
        if not excel_df.empty and col in excel_df.columns:
            uniques = sorted(set(str(v) for v in excel_df[col].dropna().unique()))
            vals.extend(uniques)
        lb = tk.Listbox(excel_headers_frame, selectmode="extended", exportselection=False, height=4, width=36)
        for v in vals:
            lb.insert(tk.END, v)
        lb.grid(row=idx, column=3, sticky="w", padx=4, pady=2)
        excel_filter_controls.append((col, var, lb))
    excel_headers_frame.columnconfigure(1, weight=1)
    excel_headers_frame.columnconfigure(3, weight=1)
    refresh_excel_filtered_tree()


def save_excel_mapping_from_controls():
    """Persist selections from the mapping UI."""
    global excel_mapping
    if not excel_mapping_controls:
        return
    # Ensure only one Spirit Code column
    first_spirit = None
    raw_map = {}
    for col, var in excel_mapping_controls:
        choice = var.get()
        if choice == "Spirit Code":
            if first_spirit is None:
                first_spirit = col
                raw_map[col] = "Spirit Code"
            else:
                raw_map[col] = "Skip"
        else:
            raw_map[col] = choice
    excel_mapping = raw_map


def get_excel_spirit_col() -> str:
    for col, mode in excel_mapping.items():
        if mode == "Spirit Code":
            return col
    return ""


def get_excel_body_cols() -> list[str]:
    return [col for col, mode in excel_mapping.items() if mode == "Include in Body"]


def excel_compute_matches():
    """Build cached matches between Excel rows and main df with availability info."""
    global excel_filtered_cache
    excel_filtered_cache = []
    spirit_col = get_excel_spirit_col()
    if not spirit_col or excel_df.empty or df.empty:
        return
    body_cols = get_excel_body_cols()
    idx = 0
    for _, erow in excel_df.iterrows():
        spirit = str(erow.get(spirit_col, "")).strip()
        if not spirit:
            continue
        df_match = df[df["Spirit Code"].astype(str) == spirit]
        if df_match.empty:
            continue
        drow = df_match.iloc[0]
        availability = {}
        recipients_all = []
        for role, var in role_send_vars.items():
            if var.get() == "Skip":
                availability[role] = "No"
                continue
            addrs = get_role_addresses(drow, role)
            if addrs:
                availability[role] = "Yes"
                recipients_all.extend(addrs)
            else:
                availability[role] = "No"
        excel_filtered_cache.append(
            {
                "id": str(idx),
                "spirit": spirit,
                "hotel": drow.get("Hotel", ""),
                "recips": "; ".join([r for r in recipients_all if r]),
                "availability": availability,
                "erow": erow,
                "drow": drow,
                "body_cols": body_cols,
            }
        )
        idx += 1


def refresh_excel_filtered_tree():
    """Populate the Excel filtered list."""
    if excel_filtered_tree is None:
        return
    save_excel_mapping_from_controls()
    excel_compute_matches()
    # apply filters per column
    filtered = excel_filtered_cache
    active_filters = []
    for col, _map_var, lb in excel_filter_controls:
        choices = [lb.get(i) for i in lb.curselection()] if lb is not None else []
        if choices:
            active_filters.append((col, choices))
    if active_filters:
        temp = []
        for m in filtered:
            erow = m["erow"]
            ok = True
            for col, crit_list in active_filters:
                val = erow.get(col, "")
                if pd.isna(val):
                    val_str = ""
                else:
                    val_str = str(val)
                match_ok = False
                for crit in crit_list:
                    if crit == "<blank>" and val_str.strip() == "":
                        match_ok = True
                        break
                    if val_str == crit:
                        match_ok = True
                        break
                if not match_ok:
                    ok = False
                    break
            if ok:
                temp.append(m)
        filtered = temp
    # fill tree
    for item in excel_filtered_tree.get_children():
        excel_filtered_tree.delete(item)
    for match in filtered:
        availability = match["availability"]
        def mark(role_key):
            return "✓" if availability.get(role_key, "No") == "Yes" else "✗"
        excel_filtered_tree.insert(
            "",
            "end",
            iid=match["id"],
            values=(
                match["spirit"],
                match["hotel"],
                match["recips"],
                mark("AVP"),
                mark("MD"),
                mark("GM"),
                mark("Engineering"),
                mark("DOF"),
                mark("RegionalEngineeringSpecialist"),
            ),
        )
    # update summary
    if excel_filter_summary_var is not None:
        summaries = []
        for col, _map_var, lb in excel_filter_controls:
            choices = [lb.get(i) for i in lb.curselection()] if lb is not None else []
            if choices:
                summaries.append(f"{col}: {', '.join(choices)}")
        excel_filter_summary_var.set("; ".join(summaries))
    if excel_filtered_count_var is not None:
        excel_filtered_count_var.set(f"Gefiltert: {len(filtered)}")
    refresh_excel_selected_tree()


def refresh_excel_selected_tree():
    """Populate the Excel selected list."""
    if excel_selected_tree is None:
        return
    for item in excel_selected_tree.get_children():
        excel_selected_tree.delete(item)
    for key, match in excel_selected_cache.items():
        availability = match["availability"]
        def mark(role_key):
            return "✓" if availability.get(role_key, "No") == "Yes" else "✗"
        excel_selected_tree.insert(
            "",
            "end",
            iid=key,
            values=(
                match["spirit"],
                match["hotel"],
                match["recips"],
                mark("AVP"),
                mark("MD"),
                mark("GM"),
                mark("Engineering"),
                mark("DOF"),
                mark("RegionalEngineeringSpecialist"),
            ),
        )


def excel_add_selected():
    """Move selected rows from filtered to selected."""
    if excel_filtered_tree is None:
        return
    chosen = excel_filtered_tree.selection()
    for cid in chosen:
        match = next((m for m in excel_filtered_cache if m["id"] == cid), None)
        if match:
            excel_selected_cache[cid] = match
    refresh_excel_selected_tree()


def excel_add_all():
    if not excel_filtered_cache:
        return
    for match in excel_filtered_cache:
        excel_selected_cache[match["id"]] = match
    refresh_excel_selected_tree()


def excel_remove_selected():
    if excel_selected_tree is None:
        return
    chosen = excel_selected_tree.selection()
    for cid in chosen:
        excel_selected_cache.pop(cid, None)
    refresh_excel_selected_tree()


def excel_clear_selected():
    excel_selected_cache.clear()
    refresh_excel_selected_tree()


def prompt_excel_compose():
    """Open dialog to compose emails from Excel rows."""
    save_excel_mapping_from_controls()
    if not excel_selected_cache:
        messagebox.showinfo(
            "Keine Auswahl",
            "Bitte zuerst Hotels in die Liste 'Ausgewaehlte Excel-Hotels' uebernehmen und dann erneut 'Emails erstellen' klicken.",
        )
        return
    if excel_df.empty:
        messagebox.showinfo("Excel Emails", "Bitte zunaechst eine Excel-Liste laden.")
        return
    spirit_col = get_excel_spirit_col()
    if not spirit_col:
        messagebox.showinfo("Excel Emails", "Bitte eine Spalte als Spirit Code festlegen.")
        return
    matches = []
    for _, erow in excel_df.iterrows():
        spirit = str(erow.get(spirit_col, "")).strip()
        if not spirit:
            continue
        df_match = df[df["Spirit Code"].astype(str) == spirit]
        if not df_match.empty:
            matches.append((erow, df_match.iloc[0]))
    if not matches:
        messagebox.showinfo("Excel Emails", "Keine passenden Spirit Codes im Hauptdatensatz gefunden.")
        return

    # Respect manual selection (if any)
    if excel_selected_cache:
        matches = [(v["erow"], v["drow"]) for v in excel_selected_cache.values()]

    dialog = tk.Toplevel(root)
    dialog.title("Excel basierte Emails")
    dialog.geometry("980x740")

    ttk.Label(dialog, text="Betreff (Platzhalter erlaubt):").pack(anchor="w", padx=8, pady=(8, 2))
    subj_var = tk.StringVar(value="Information fuer {hotel}")
    subj_entry = ttk.Entry(dialog, textvariable=subj_var)
    subj_entry.pack(fill="x", padx=8)

    ttk.Label(dialog, text="Notiz / Einleitung (HTML, Platzhalter erlaubt):").pack(anchor="w", padx=8, pady=(8, 2))
    body_text = tk.Text(dialog, height=10, font=("Aptos", 12))
    body_text.pack(fill="both", expand=True, padx=8)

    link_frame = ttk.Frame(dialog)
    link_frame.pack(fill="x", padx=8, pady=(0, 4))
    ttk.Button(link_frame, text="Link einfuegen...", command=lambda: open_link_dialog(body_text)).pack(side="left")

    ph_frame = ttk.Frame(dialog)
    ph_frame.pack(fill="x", padx=8, pady=4)
    ttk.Label(ph_frame, text="Platzhalter:").pack(side="left")
    ph_var = tk.StringVar(value=PLACEHOLDERS[0])
    ph_combo = ttk.Combobox(ph_frame, textvariable=ph_var, values=PLACEHOLDERS, state="readonly", width=20)
    ph_combo.pack(side="left", padx=4)
    ttk.Button(ph_frame, text="In Text einfuegen", command=lambda: body_text.insert("insert", ph_var.get())).pack(side="left", padx=4)
    ttk.Button(ph_frame, text="In Betreff einfuegen", command=lambda: subj_entry.insert("insert", ph_var.get())).pack(side="left", padx=4)

    sigs = load_signatures()
    ttk.Label(dialog, text="Signatur:").pack(anchor="w", padx=8, pady=(6, 2))
    sig_var = tk.StringVar(value="None")
    sig_combo = ttk.Combobox(dialog, textvariable=sig_var, values=list(sigs.keys()), state="readonly")
    sig_combo.pack(fill="x", padx=8)
    if forward_template.get("body_text", ""):
        sig_var.set("None")
        sig_combo.state(["disabled"])
        ttk.Label(dialog, text="Signatur deaktiviert, da eine weitergeleitete Email eingebettet wird.", foreground="gray").pack(anchor="w", padx=8, pady=(2, 2))

    def do_send():
        note_text = body_text.get("1.0", "end").strip()
        signature_text = sigs.get(sig_var.get(), "")
        subj_tpl = subj_var.get().strip() or "Information"
        body_cols = get_excel_body_cols()
        mode = excel_mode_var.get() or "dedicated"
        send_excel_emails(matches, subj_tpl, note_text, signature_text, body_cols, mode)
        dialog.destroy()

    btns = ttk.Frame(dialog)
    btns.pack(fill="x", padx=8, pady=10)
    ttk.Button(btns, text="Abbrechen", command=dialog.destroy).pack(side="right", padx=4)
    ttk.Button(btns, text="Emails erstellen", command=do_send).pack(side="right", padx=4)


def send_excel_emails(matches, subj_tpl: str, note_text: str, signature_text: str, body_cols: list[str], mode: str):
    """Create Outlook drafts from Excel rows."""
    if os.name != "nt":
        messagebox.showerror("Unsupported Platform", "Outlook email drafting ist nur unter Windows verfuegbar.")
        return
    if not WIN32COM_AVAILABLE:
        messagebox.showerror("Outlook Not Available", "Dieses Feature erfordert Outlook und pywin32.")
        return
    try:
        outlook = get_outlook_app()
    except Exception as exc:
        messagebox.showerror("Outlook Fehler", f"Outlook konnte nicht gestartet werden: {exc}")
        return

    attach_enabled = attachments_enabled_var.get() if attachments_enabled_var is not None else False
    attach_root = attachments_root_var.get() if attachments_root_var is not None else ""

    def note_to_html_block(text: str) -> str:
        """Convert note text to HTML; preserve existing HTML/anchors."""
        if not text:
            return ""
        lower = text.lower()
        # If user already provided HTML (e.g., <a ...>), trust it
        if "<a " in lower or "<table" in lower or "<div" in lower or "<p" in lower:
            return text.replace("\n", "<br>")
        # Otherwise, convert markdown links and escape safely
        rendered = render_with_signature(text, {"html": "", "text": ""}, False, "", False)
        if rendered.get("html"):
            html_block = rendered["html"]
            l2 = html_block.lower()
            body_start = l2.find("<body")
            if body_start != -1:
                gt = l2.find(">", body_start)
                end_tag = l2.rfind("</body>")
                if gt != -1 and end_tag != -1 and end_tag > gt:
                    return html_block[gt + 1 : end_tag]
            return html_block
        return html.escape(text).replace("\n", "<br>")

    def render_body(erow: pd.Series) -> str:
        note_html_local = note_to_html_block(note_text) if note_text else ""
        table_html = html_table_from_excel_row(erow, body_cols)
        if note_html_local and table_html:
            return note_html_local + "<br><br>" + table_html
        return note_html_local + table_html

    def new_mail_item():
        nonlocal outlook
        try:
            return outlook.CreateItem(0)
        except Exception:
            try:
                outlook = get_outlook_app(force_refresh=True)
                return outlook.CreateItem(0)
            except Exception as exc:
                raise exc

    if mode == "collective":
        to_set, cc_set, bcc_set = set(), set(), set()
        combined_sections = []
        for erow, drow in matches:
            spirit = str(drow.get("Spirit Code", "")).strip()
            hotel_name = drow.get("Hotel", "")
            for role, var in role_send_vars.items():
                if var.get() == "Skip":
                    continue
                emails = get_role_addresses(drow, role)
                if var.get() == "To":
                    to_set.update(emails)
                elif var.get() == "CC":
                    cc_set.update(emails)
                elif var.get() == "BCC":
                    bcc_set.update(emails)
            section = f"<h3>{html.escape(spirit)} - {html.escape(str(hotel_name))}</h3>" + render_body(erow)
            combined_sections.append(section)

        all_recips = [r for r in list(to_set | cc_set | bcc_set) if r]
        if not all_recips:
            messagebox.showinfo("Keine Empfaenger", "Keine Empfaenger fuer Sammelmail gefunden.")
            return
        try:
            mail_item = new_mail_item()
            try:
                mail_item.BodyFormat = 2  # HTML
            except Exception:
                pass
            mail_item.To = ";".join(to_set)
            mail_item.CC = ";".join(cc_set)
            mail_item.BCC = ";".join(bcc_set)
            mail_item.Subject = subj_tpl
            rendered = render_with_signature(
                "<br>".join(combined_sections),
                {"html": "", "text": ""} if forward_template.get("body_text", "") else signature_text,
                True,
                forward_template.get("body_text", ""),
                forward_template.get("is_html", False),
            )
            if rendered.get("html"):
                mail_item.HTMLBody = rendered["html"]
            else:
                mail_item.Body = rendered.get("text", "")
            for path in forward_template.get("attachments", []):
                if os.path.isfile(path):
                    try:
                        mail_item.Attachments.Add(path)
                    except Exception:
                        pass
            if attach_enabled:
                attach_files_for_hotel(mail_item, attach_root, "")
            mail_item.Display()
        except Exception as exc:
            messagebox.showerror("Email Fehler", f"Sammelmail konnte nicht erstellt werden: {exc}")
        return

    # Dedicated per hotel
    missing_addresses = []
    for erow, drow in matches:
        to_list, cc_list, bcc_list = [], [], []
        for role, var in role_send_vars.items():
            mode_val = var.get()
            if mode_val == "Skip":
                continue
            emails = get_role_addresses(drow, role)
            if mode_val == "To":
                to_list.extend(emails)
            elif mode_val == "CC":
                cc_list.extend(emails)
            elif mode_val == "BCC":
                bcc_list.extend(emails)
        all_recips = [r for r in to_list + cc_list + bcc_list if r]
        if not all_recips:
            missing_addresses.append(drow.get("Hotel", "N/A"))
            continue
        try:
            mail_item = new_mail_item()
            try:
                mail_item.BodyFormat = 2  # HTML
            except Exception:
                pass
            mail_item.To = ";".join(to_list)
            mail_item.CC = ";".join(cc_list)
            mail_item.BCC = ";".join(bcc_list)
            mail_item.Subject = render_template(drow, subj_tpl)
            body_html = render_template(drow, render_body(erow))
            rendered = render_with_signature(
                body_html,
                {"html": "", "text": ""} if forward_template.get("body_text", "") else signature_text,
                True,
                forward_template.get("body_text", ""),
                forward_template.get("is_html", False),
            )
            if rendered.get("html"):
                mail_item.HTMLBody = rendered["html"]
            else:
                mail_item.Body = rendered.get("text", "")
            for path in forward_template.get("attachments", []):
                if os.path.isfile(path):
                    try:
                        mail_item.Attachments.Add(path)
                    except Exception:
                        pass
            if attach_enabled:
                spirit = str(drow.get("Spirit Code", "")).strip()
                attach_files_for_hotel(mail_item, attach_root, spirit)
            mail_item.Display()
        except Exception as exc:
            messagebox.showerror("Email Fehler", f"Mail fuer {drow.get('Hotel','Hotel')} konnte nicht erstellt werden: {exc}")
            return
    if missing_addresses:
        messagebox.showinfo(
            "Keine Empfaenger",
            "Fuer folgende Hotels wurden keine Empfaenger gefunden:\n" + "\n".join(missing_addresses),
        )


def draft_emails_for_selection():
    """Create Outlook draft emails for the selected hotels and roles."""
    if not selected_rows:
        messagebox.showinfo("Keine Hotels", "Bitte waehlen Sie mindestens ein Hotel aus der Auswahl aus.")
        return

    chosen_roles = [role for role, var in role_send_vars.items() if var.get() != "Skip"]

    if not chosen_roles:
        messagebox.showinfo("Keine Rollen", "Bitte waehlen Sie mindestens eine Empfaengerrolle.")
        return

    if os.name != "nt":
        messagebox.showerror("Unsupported Platform", "Outlook email drafting is only available on Windows.")
        return

    if not WIN32COM_AVAILABLE:
        messagebox.showerror(
            "Outlook Not Available",
            "This feature requires Microsoft Outlook and the 'pywin32' package (win32com.client).\nInstall with: pip install pywin32",
        )
        return

    try:
        outlook = get_outlook_app()
        mail_test = outlook.CreateItem(0)
    except Exception:
        try:
            outlook = get_outlook_app(force_refresh=True)
            mail_test = outlook.CreateItem(0)
        except Exception as exc:  # pragma: no cover - Outlook automation is Windows-specific
            messagebox.showerror("Email Error", f"Could not draft email in Outlook: {exc}")
            return

    # Prompt for subject/body with placeholders
    def open_message_dialog():
        dialog = tk.Toplevel(root)
        dialog.title("Compose Email Template")
        dialog.geometry("960x720")
        ttk.Label(dialog, text="Subject (supports placeholders):").pack(anchor="w", padx=8, pady=(8, 2))
        default_subj = forward_template["subject"] or "Hotel Information for {hotel}"
        subj_var = tk.StringVar(value=default_subj)
        subj_entry = ttk.Entry(dialog, textvariable=subj_var)
        subj_entry.pack(fill="x", padx=8)

        ttk.Label(dialog, text="Your note (appears above signature and the forwarded email):").pack(anchor="w", padx=8, pady=(8, 2))
        body_text = tk.Text(dialog, height=10, font=("Aptos", 12))
        body_text.pack(fill="both", expand=True, padx=8)
        body_text.insert("1.0", "Your message here.")

        link_frame = ttk.Frame(dialog)
        link_frame.pack(fill="x", padx=8, pady=(0, 4))
        ttk.Button(link_frame, text="Insert Link...", command=lambda: open_link_dialog(body_text)).pack(side="left")

        # Placeholder helper
        ph_frame = ttk.Frame(dialog)
        ph_frame.pack(fill="x", padx=8, pady=4)
        ttk.Label(ph_frame, text="Placeholders:").pack(side="left")
        ph_var = tk.StringVar(value=PLACEHOLDERS[0])
        ph_combo = ttk.Combobox(ph_frame, textvariable=ph_var, values=PLACEHOLDERS, state="readonly", width=20)
        ph_combo.pack(side="left", padx=4)
        def insert_placeholder(target="body"):
            ph = ph_var.get()
            if target == "body":
                body_text.insert("insert", ph)
            else:
                subj_entry.insert("insert", ph)
        ttk.Button(ph_frame, text="Insert in Body", command=lambda: insert_placeholder("body")).pack(side="left", padx=4)
        ttk.Button(ph_frame, text="Insert in Subject", command=lambda: insert_placeholder("subj")).pack(side="left", padx=4)

        placeholder_text = (
            "Placeholders:\n"
            "{hotel}, {spirit_code}, {city}, {relationship}, {brand}, {brand_band}, "
            "{region}, {country}, {owner}, {rooms}\n"
            "They will be replaced per hotel."
        )
        ttk.Label(dialog, text=placeholder_text, foreground="gray").pack(anchor="w", padx=8, pady=(4, 8))

        sigs = load_signatures()
        ttk.Label(dialog, text="Signature:").pack(anchor="w", padx=8, pady=(4, 2))
        sig_var = tk.StringVar(value="None")
        sig_combo = ttk.Combobox(dialog, textvariable=sig_var, values=list(sigs.keys()), state="readonly")
        sig_combo.pack(fill="x", padx=8, pady=(0, 6))
        if forward_template.get("body_text"):
            sig_var.set("None")
            sig_combo.state(["disabled"])
            ttk.Label(
                dialog,
                text="Signature disabled for forwarded emails to keep the original body intact.",
                foreground="gray",
            ).pack(anchor="w", padx=8, pady=(2, 6))

        def render_and_send():
            subject_template = subj_var.get()
            body_template = body_text.get("1.0", "end")
            dialog.destroy()
            sig_choice = sigs.get(sig_var.get(), {"html": "", "text": ""})
            if forward_template.get("body_text"):
                sig_choice = {"html": "", "text": ""}
            send_emails(
                subject_template,
                body_template,
                sig_choice,
                body_is_html=False,
                forward_html=forward_template.get("body_text", ""),
                forward_is_html=forward_template.get("is_html", False),
            )
        ttk.Button(dialog, text="Create Drafts", command=render_and_send).pack(pady=6)

    def send_emails(
        subject_template: str,
        body_template: str,
        signature_text: dict,
        body_is_html: bool = False,
        forward_html: str = "",
        forward_is_html: bool = False,
    ):
        missing_addresses = []
        attach_enabled = attachments_enabled_var.get() if attachments_enabled_var else False
        attach_root = attachments_root_var.get() if attachments_root_var else ""
        # If forwarding, ignore signatures to prevent Outlook from stripping the forwarded body
        if forward_html:
            signature_text = {"html": "", "text": ""}

        for row_idx, row in selected_rows.items():
            to_list = []
            cc_list = []
            bcc_list = []

            for role in chosen_roles:
                emails = get_role_addresses(row, role)
                mode = role_send_vars.get(role).get() if role in role_send_vars else "To"
                if mode == "To":
                    to_list.extend(emails)
                elif mode == "CC":
                    cc_list.extend(emails)
                elif mode == "BCC":
                    bcc_list.extend(emails)

            all_recips = [r for r in to_list + cc_list + bcc_list if r]
            if not all_recips:
                missing_addresses.append(row.get("Hotel", "N/A"))
                continue

            try:
                mail_item = outlook.CreateItem(0)
                try:
                    mail_item.BodyFormat = 2  # olFormatHTML
                except Exception:
                    pass
                mail_item.To = ";".join(to_list)
                mail_item.CC = ";".join(cc_list)
                mail_item.BCC = ";".join(bcc_list)
                hotel_name = row.get("Hotel", "Hotel")
                mail_item.Subject = render_template(row, subject_template)
                rendered = render_with_signature(
                    render_template(row, body_template),
                    signature_text,
                    body_is_html,
                    forward_template.get("body_text", ""),
                    forward_template.get("is_html", False),
                )
                if rendered.get("html"):
                    mail_item.HTMLBody = rendered["html"]
                else:
                    mail_item.Body = rendered.get("text", "")
                # Attach captured forward attachments first
                for path in forward_template.get("attachments", []):
                    if os.path.isfile(path):
                        try:
                            mail_item.Attachments.Add(path)
                        except Exception:
                            pass
                # Attach files if enabled
                if attach_enabled:
                    attach_files_for_hotel(mail_item, attach_root, str(row.get("Spirit Code", "")).strip())
                mail_item.Display()
            except Exception as exc:
                messagebox.showerror("Email Error", f"Could not draft email for {row.get('Hotel', 'Hotel')}: {exc}")
                return

        if missing_addresses:
            messagebox.showinfo(
                "Keine Empfaenger",
                "Fuer folgende Hotels wurden keine E-Mail-Adressen in den gewaehlten Rollen gefunden:\n" + "\n".join(missing_addresses),
            )

    open_message_dialog()


def draft_collective_email():
    """Create a single Outlook draft to all recipients from selected hotels (multi-email tab)."""
    if not selected_rows:
        messagebox.showinfo("Keine Hotels", "Bitte waehlen Sie mindestens ein Hotel aus der Auswahl aus.")
        return
    if os.name != "nt":
        messagebox.showerror("Unsupported Platform", "Outlook email drafting is only available on Windows.")
        return
    if not WIN32COM_AVAILABLE:
        messagebox.showerror(
            "Outlook Not Available",
            "This feature requires Microsoft Outlook and the 'pywin32' package (win32com.client).\nInstall with: pip install pywin32",
        )
        return

    dialog = tk.Toplevel(root)
    dialog.title("Collective Email")
    dialog.geometry("700x550")
    ttk.Label(dialog, text="Subject:").pack(anchor="w", padx=8, pady=(8, 2))
    subj_var = tk.StringVar(value=forward_template.get("subject", ""))
    subj_entry = ttk.Entry(dialog, textvariable=subj_var)
    subj_entry.pack(fill="x", padx=8)

    ttk.Label(dialog, text="Body (plain text; note + signature above forwarded content if any):").pack(anchor="w", padx=8, pady=(8, 2))
    body_text = tk.Text(dialog, height=10)
    body_text.pack(fill="both", expand=True, padx=8)
    body_text.insert("1.0", "Your message here.")

    link_frame = ttk.Frame(dialog)
    link_frame.pack(fill="x", padx=8, pady=(2, 4))
    ttk.Button(link_frame, text="Insert Link...", command=lambda: open_link_dialog(body_text)).pack(side="left")

    sigs = load_signatures()
    ttk.Label(dialog, text="Signature:").pack(anchor="w", padx=8, pady=(4, 2))
    sig_var = tk.StringVar(value="None")
    ttk.Combobox(dialog, textvariable=sig_var, values=list(sigs.keys()), state="readonly").pack(fill="x", padx=8, pady=(0, 6))

    attach_enabled = attachments_enabled_var.get() if attachments_enabled_var else False
    attach_root = attachments_root_var.get() if attachments_root_var else ""

    def render_and_send():
        subject = subj_var.get()
        body = body_text.get("1.0", "end")
        dialog.destroy()

        to_list, cc_list, bcc_list = [], [], []
        for row_idx, row in selected_rows.items():
            for role in role_send_vars:
                mode = role_send_vars[role].get()
                if mode == "Skip":
                    continue
                emails = get_role_addresses(row, role)
                for em in emails:
                    if mode == "To":
                        to_list.append(em)
                    elif mode == "CC":
                        cc_list.append(em)
                    elif mode == "BCC":
                        bcc_list.append(em)

        def dedup(seq):
            seen = set()
            out = []
            for x in seq:
                key = x.strip().lower()
                if key and key not in seen:
                    seen.add(key)
                    out.append(x.strip())
            return out

        to_list[:] = dedup(to_list)
        cc_list[:] = dedup(cc_list)
        bcc_list[:] = dedup(bcc_list)

        if not (to_list or cc_list or bcc_list):
            messagebox.showinfo("No Recipients", "No email addresses selected across roles.")
            return

        try:
            outlook = get_outlook_app()
            mail_item = outlook.CreateItem(0)
            try:
                mail_item.BodyFormat = 2  # olFormatHTML
            except Exception:
                pass
        except Exception as exc:
            messagebox.showerror("Email Error", f"Could not draft email in Outlook: {exc}")
            return

        mail_item.To = ";".join(to_list)
        mail_item.CC = ";".join(cc_list)
        mail_item.BCC = ";".join(bcc_list)
        mail_item.Subject = subject

        sig_entry = sigs.get(sig_var.get(), {"html": "", "text": ""})
        rendered = render_with_signature(
            body,
            sig_entry,
            False,
            forward_template.get("body_text", ""),
            forward_template.get("is_html", False),
        )
        if rendered.get("html"):
            mail_item.HTMLBody = rendered["html"]
        else:
            mail_item.Body = rendered.get("text", "")

        for path in forward_template.get("attachments", []):
            if os.path.isfile(path):
                try:
                    mail_item.Attachments.Add(path)
                except Exception:
                    pass

        if attach_enabled and attach_root:
            common_dir = os.path.join(attach_root, attachments_common_dir)
            common_files = glob.glob(os.path.join(common_dir, "*")) if os.path.isdir(common_dir) else []
            for path in common_files:
                if os.path.isfile(path):
                    try:
                        mail_item.Attachments.Add(path)
                    except Exception:
                        pass
            spirit_paths = set()
            for row in selected_rows.values():
                spirit_code = str(row.get("Spirit Code", "")).strip()
                if not spirit_code:
                    continue
                for path in collect_spirit_dirs(attach_root, spirit_code):
                    if os.path.isfile(path):
                        spirit_paths.add(path)
            for path in spirit_paths:
                try:
                    mail_item.Attachments.Add(path)
                except Exception:
                    pass

        mail_item.Display()

    ttk.Button(dialog, text="Create Collective Draft", command=render_and_send).pack(pady=6)


# ---------------------------------------------------------------------------
# Lookup (single hotel) helpers
# ---------------------------------------------------------------------------
def lookup(spirit_entry, hotel_var_local):
    if df.empty:
        messagebox.showwarning(
            "Keine Daten",
            "Es sind derzeit keine Daten geladen. Bitte laden Sie eine Excel-Datei ueber 'Datei -> Datendatei oeffnen'.",
        )
        clear_detail_panel("No data loaded.")
        return

    spirit = spirit_entry.get().strip()
    hotel = hotel_var_local.get().strip()
    city_col = get_city_col()

    if spirit:
        result = df[df["Spirit Code"].astype(str).str.lower() == spirit.lower()]
    elif hotel:
        mask = df["Hotel"].astype(str).str.contains(hotel, case=False, na=False)
        if city_col and city_col in df.columns:
            mask |= df[city_col].astype(str).str.contains(hotel, case=False, na=False)
        result = df[mask]
    else:
        messagebox.showwarning("Whoops", "Enter Spirit Code or pick a hotel.")
        clear_detail_panel("Enter Spirit Code or hotel to view details.")
        return

    if result.empty:
        messagebox.showinfo("Nada", "No matching hotel found.")
        clear_detail_panel("No matching hotel found.")
        return

    if len(result) == 1:
        row = result.iloc[0]
        populate_detail_panel(row)
    else:
        show_search_results(result)


def show_details_gui(row):
    global detail_row_current
    detail_row_current = row
    win = tk.Toplevel(root)
    win.title(f"Details for {row.get('Hotel', 'N/A')}")
    win.geometry("700x760")
    win.minsize(500, 400)

    info_frame = ttk.LabelFrame(win, text="Hotel Information", padding="10")
    info_frame.pack(padx=10, pady=10, fill="x")

    general_info = [
        ("Spirit Code", "Spirit Code"),
        ("Hotel", "Hotel"),
        ("City", get_city_col()),
        ("Country/Area", get_country_col() or get_country_fallback_col()),
        ("Relationship", "Relationship"),
        ("Rooms", "Rooms"),
        ("JV", "JV"),
        ("JV Percent", "JV Percent"),
        ("Owner", "Owner"),
    ]

    row_idx_info = 0
    for label_text, col in general_info:
        if col and col in row and pd.notna(row[col]):
            tk.Label(info_frame, text=f"{label_text}:", anchor="w", font=("Arial", 10, "bold")).grid(
                row=row_idx_info, column=0, sticky="w", padx=5, pady=2
            )
            tk.Label(info_frame, text=row[col], anchor="w", font=("Arial", 10)).grid(
                row=row_idx_info, column=1, sticky="w", padx=5, pady=2
            )
            row_idx_info += 1

    roles_frame = ttk.LabelFrame(win, text="Key Personnel (Select for Email)", padding="10")
    roles_frame.pack(padx=10, pady=10, fill="both", expand=False)

    roles_to_checkbox = {}
    if get_avp_col():
        roles_to_checkbox["AVP"] = get_avp_col()
    if get_md_col():
        roles_to_checkbox["MD"] = get_md_col()
    if get_gm_col():
        roles_to_checkbox["GM"] = get_gm_col()
    if get_eng_col():
        roles_to_checkbox["Engineering"] = get_eng_col()
    if get_dof_col():
        roles_to_checkbox["DOF"] = get_dof_col()
    if get_reg_eng_spec_col():
        roles_to_checkbox["Regional Eng Specialist"] = get_reg_eng_spec_col()

    checkbox_vars = []
    row_idx = 0
    if not roles_to_checkbox:
        tk.Label(roles_frame, text="No role columns configured.", anchor="w", foreground="gray").grid(
            row=row_idx, column=0, sticky="w", padx=5, pady=2
        )
    else:
        for role, email_col in roles_to_checkbox.items():
            email_address = row.get(email_col)
            if email_col in row.index and pd.notna(email_address):
                var = tk.BooleanVar()
                chk = ttk.Checkbutton(roles_frame, text=f"{role}: {email_address}", variable=var)
                chk.grid(row=row_idx, column=0, sticky="w", padx=5, pady=2)
                canonical_role = "RegionalEngineeringSpecialist" if role.startswith("Regional") else role
                checkbox_vars.append((var, str(email_address), canonical_role))
                row_idx += 1
            else:
                tk.Label(roles_frame, text=f"{role}: N/A (Email not found)", anchor="w", foreground="gray").grid(
                    row=row_idx, column=0, sticky="w", padx=5, pady=2
                )
                row_idx += 1

    button_frame = ttk.Frame(win)
    button_frame.pack(pady=10)

    tk.Button(button_frame, text="Close", command=win.destroy).pack(side="left", padx=10)
    tk.Button(
        button_frame,
        text="Start Email",
        command=lambda: draft_email_single(checkbox_vars, row.get("Hotel", "N/A"), win),
    ).pack(side="left", padx=10)


def draft_email_single(checkbox_vars, hotel_name, details_window=None):
    if os.name != "nt":
        messagebox.showerror("Unsupported Platform", "Outlook email drafting is only available on Windows.")
        return

    if not WIN32COM_AVAILABLE:
        messagebox.showerror(
            "Outlook Not Available",
            "This feature requires Microsoft Outlook and the 'pywin32' package (win32com.client).\nInstall with: pip install pywin32",
        )
        return

    # Compose dialog for single email
    def open_single_template():
        dialog = tk.Toplevel(root)
        dialog.title("Compose Email Template (Single Hotel)")
        dialog.geometry("960x640")
        ttk.Label(dialog, text="Subject (supports placeholders):").pack(anchor="w", padx=8, pady=(8, 2))
        subj_var = tk.StringVar(value="Hotel Information for {hotel}")
        subj_entry = ttk.Entry(dialog, textvariable=subj_var)
        subj_entry.pack(fill="x", padx=8)

        ttk.Label(dialog, text="Body (supports placeholders):").pack(anchor="w", padx=8, pady=(8, 2))
        body_text = tk.Text(dialog, height=8)
        body_text.pack(fill="both", expand=True, padx=8)
        body_text.insert(
            "1.0",
            "Hotel: {hotel}\nSpirit: {spirit_code}\nCity: {city}\nBrand: {brand}\n\nYour message here.",
        )

        link_frame = ttk.Frame(dialog)
        link_frame.pack(fill="x", padx=8, pady=(0, 4))
        ttk.Button(link_frame, text="Insert Link...", command=lambda: open_link_dialog(body_text)).pack(side="left")

        # Recipients with per-recipient To/CC/BCC (independent of Setup tab)
        rec_frame = ttk.LabelFrame(dialog, text="Recipients", padding=6)
        rec_frame.pack(fill="x", padx=8, pady=6)
        recipient_controls = []
        row_idx = 0
        for orig_var, email, role_key in checkbox_vars:
            local_sel = tk.BooleanVar(value=orig_var.get())
            local_mode = tk.StringVar(value="To")
            ttk.Checkbutton(rec_frame, text=f"{role_key}: {email}", variable=local_sel).grid(row=row_idx, column=0, sticky="w", padx=4, pady=2)
            ttk.Combobox(rec_frame, textvariable=local_mode, values=["To", "CC", "BCC"], state="readonly", width=6).grid(row=row_idx, column=1, sticky="w", padx=4, pady=2)
            recipient_controls.append((local_sel, local_mode, email))
            row_idx += 1

        # Placeholder helper
        ph_frame = ttk.Frame(dialog)
        ph_frame.pack(fill="x", padx=8, pady=4)
        ttk.Label(ph_frame, text="Placeholders:").pack(side="left")
        ph_var = tk.StringVar(value=PLACEHOLDERS[0])
        ph_combo = ttk.Combobox(ph_frame, textvariable=ph_var, values=PLACEHOLDERS, state="readonly", width=20)
        ph_combo.pack(side="left", padx=4)
        def insert_placeholder(target="body"):
            ph = ph_var.get()
            if target == "body":
                body_text.insert("insert", ph)
            else:
                subj_entry.insert("insert", ph)
        ttk.Button(ph_frame, text="Insert in Body", command=lambda: insert_placeholder("body")).pack(side="left", padx=4)
        ttk.Button(ph_frame, text="Insert in Subject", command=lambda: insert_placeholder("subj")).pack(side="left", padx=4)

        placeholder_text = (
            "Placeholders: {hotel}, {spirit_code}, {city}, {relationship}, {brand}, {brand_band}, "
            "{region}, {country}, {owner}, {rooms}"
        )
        ttk.Label(dialog, text=placeholder_text, foreground="gray").pack(anchor="w", padx=8, pady=(4, 8))

        sigs = load_signatures()
        ttk.Label(dialog, text="Signature:").pack(anchor="w", padx=8, pady=(4, 2))
        sig_var = tk.StringVar(value="None")
        sig_combo = ttk.Combobox(dialog, textvariable=sig_var, values=list(sigs.keys()), state="readonly")
        sig_combo.pack(fill="x", padx=8, pady=(0, 6))

        def send_single():
            subject_template = subj_var.get()
            body_template = body_text.get("1.0", "end").rstrip("\n")
            dialog.destroy()

            try:
                outlook = get_outlook_app()
                mail_item = outlook.CreateItem(0)
                try:
                    mail_item.BodyFormat = 2  # olFormatHTML
                except Exception:
                    pass
            except Exception:
                try:
                    outlook = get_outlook_app(force_refresh=True)
                    mail_item = outlook.CreateItem(0)
                except Exception as exc:  # pragma: no cover - Outlook automation is Windows-specific
                    messagebox.showerror("Email Error", f"Could not draft email in Outlook: {exc}")
                    return

            to_list, cc_list, bcc_list = [], [], []
            for var, email, role_key in checkbox_vars:
                if var.get() and email:
                    emails = normalize_emails(email)
                    mode = role_send_vars.get(role_key).get() if role_key in role_send_vars else "To"
                    if mode == "To":
                        to_list.extend(emails)
                    elif mode == "CC":
                        cc_list.extend(emails)
                    elif mode == "BCC":
                        bcc_list.extend(emails)

            all_recips = [r for r in to_list + cc_list + bcc_list if r]
            if not all_recips:
                messagebox.showinfo("No Recipients", "No email addresses selected.")
                return

            mail_item.To = ";".join(to_list)
            mail_item.CC = ";".join(cc_list)
            mail_item.BCC = ";".join(bcc_list)

            mail_item.Subject = render_template(detail_row_current, subject_template)
            sig_entry = sigs.get(sig_var.get(), {"html": "", "text": ""})
            rendered = render_with_signature(
                render_template(detail_row_current, body_template),
                sig_entry,
                False,
            )
            if rendered.get("html"):
                mail_item.HTMLBody = rendered["html"]
            else:
                mail_item.Body = rendered.get("text", "")
            # Apply per-recipient routing from dialog
            to_list.clear()
            cc_list.clear()
            bcc_list.clear()
            for sel_var, mode_var, email in recipient_controls:
                if sel_var.get() and email:
                    for em in normalize_emails(email):
                        mode = mode_var.get()
                        if mode == "To":
                            to_list.append(em)
                        elif mode == "CC":
                            cc_list.append(em)
                        elif mode == "BCC":
                            bcc_list.append(em)
            if not (to_list or cc_list or bcc_list):
                messagebox.showinfo("No Recipients", "No recipients selected.")
                return
            mail_item.To = ";".join(to_list)
            mail_item.CC = ";".join(cc_list)
            mail_item.BCC = ";".join(bcc_list)
            # Attach files if enabled (lookup-tab specific)
            attach_enabled = single_attachments_enabled_var.get() if single_attachments_enabled_var else False
            attach_root = single_attachments_root_var.get() if single_attachments_root_var else ""
            if attach_enabled:
                attach_files_for_hotel(mail_item, attach_root, str(detail_row_current.get("Spirit Code", "")).strip())
            mail_item.Display()
            if details_window is not None:
                details_window.destroy()

        ttk.Button(dialog, text="Create Draft", command=send_single).pack(pady=6)

    open_single_template()


def show_search_results(results_df):
    win = tk.Toplevel(root)
    win.title("Suchergebnisse")
    win.geometry("500x300")

    tree = ttk.Treeview(win, columns=("Spirit Code", "Hotel", "City"), show="headings")
    tree.heading("Spirit Code", text="Spirit Code")
    tree.heading("Hotel", text="Hotel")

    tree.heading("City", text="City")

    tree.column("Spirit Code", width=100)
    tree.column("Hotel", width=250)
    tree.column("City", width=130)
    tree.pack(fill="both", expand=True, padx=10, pady=10)

    for _, result_row in results_df.iterrows():
        tree.insert("", "end", values=(result_row.get("Spirit Code", ""), result_row.get("Hotel", ""), get_city_value(result_row)))

    def open_selected(event=None):
        selected = tree.focus()
        if not selected:
            messagebox.showinfo("Auswahl", "Bitte waehlen Sie einen Eintrag aus.")
            return
        try:
            row_index = tree.index(selected)
        except tk.TclError:
            messagebox.showerror("Fehler", "Die Auswahl konnte nicht ermittelt werden.")
            return

        if row_index >= len(results_df):
            messagebox.showerror("Fehler", "Der ausgewaehlte Eintrag konnte nicht geladen werden.")
            return

        populate_detail_panel(results_df.iloc[row_index])
        win.destroy()

    btn_frame = ttk.Frame(win)
    btn_frame.pack(pady=(0, 10))
    ttk.Button(btn_frame, text="Details anzeigen", command=open_selected).pack(side="left", padx=5)
    ttk.Button(btn_frame, text="Schliessen", command=win.destroy).pack(side="left", padx=5)

    tree.bind("<Double-1>", open_selected)


# ---------------------------------------------------------------------------
# GUI construction
# ---------------------------------------------------------------------------
root = tk.Tk()
root.title("Hotel Lookup")
root.geometry("1250x900")
ensure_style()
root.after(0, show_splash)
show_splash()

status_var = tk.StringVar(value="Lade Daten ...")
attachments_enabled_var = tk.BooleanVar(value=False)
attachments_root_var = tk.StringVar(value="")
excel_file_label_var = tk.StringVar(value="Keine Datei geladen")
excel_mode_var = tk.StringVar(value="dedicated")

# Initialize column selection vars
brand_col_var = tk.StringVar(value=DEFAULT_BRAND_COL)
region_col_var = tk.StringVar(value=DEFAULT_REGION_COL)
country_col_var = tk.StringVar(value=DEFAULT_COUNTRY_COL)
country_fallback_col_var = tk.StringVar(value=DEFAULT_COUNTRY_FALLBACK_COL)
city_col_var = tk.StringVar(value=DEFAULT_CITY_COL)
brand_band_col_var = tk.StringVar(value=DEFAULT_BRAND_BAND_COL)
relationship_col_var = tk.StringVar(value=DEFAULT_RELATIONSHIP_COL)
hyatt_date_col_var = tk.StringVar(value=DEFAULT_HYATT_DATE_COL)
gm_col_var = tk.StringVar(value=DEFAULT_GM_COL)
eng_col_var = tk.StringVar(value=DEFAULT_ENG_COL)
dof_col_var = tk.StringVar(value=DEFAULT_DOF_COL)
avp_col_var = tk.StringVar(value="AVP of Ops")
md_col_var = tk.StringVar(value="SVP / Managing Director")
reg_eng_spec_col_var = tk.StringVar(value="None")

# Menu bar
menubar = tk.Menu(root)
file_menu = tk.Menu(menubar, tearoff=0)
file_menu.add_command(label="Datendatei oeffnen", command=prompt_for_file)
file_menu.add_command(label="Konfiguration laden", command=load_config_file)
file_menu.add_command(label="Konfiguration speichern", command=save_config_file)
file_menu.add_separator()
file_menu.add_command(label="Beenden", command=root.quit)
menubar.add_cascade(label="Datei", menu=file_menu)

def reopen_splash():
    show_splash()
    update_splash(data_file_path, "Status: Ready")

help_menu = tk.Menu(menubar, tearoff=0)
help_menu.add_command(label="About / Splash", command=reopen_splash)
menubar.add_cascade(label="About", menu=help_menu)

def show_readme():
    readme_path = os.path.join(BASE_DIR, "README.md")
    content = "README.md not found."
    if os.path.isfile(readme_path):
        try:
            with open(readme_path, "r", encoding="utf-8") as fh:
                content = fh.read()
        except Exception as exc:
            content = f"Could not read README.md:\n{exc}"
    win = tk.Toplevel(root)
    win.title("Instructions (README)")
    win.geometry("760x520")
    text = tk.Text(win, wrap="word")
    text.insert("1.0", content)
    text.config(state="disabled")
    text.pack(fill="both", expand=True, padx=6, pady=6)
    scroll = ttk.Scrollbar(win, command=text.yview)
    text.configure(yscrollcommand=scroll.set)
    scroll.pack(side="right", fill="y")

help_menu2 = tk.Menu(menubar, tearoff=0)
help_menu2.add_command(label="Instructions", command=show_readme)
menubar.add_cascade(label="Help", menu=help_menu2)

root.config(menu=menubar)

notebook = ttk.Notebook(root)
notebook.pack(fill="both", expand=True)

# ---------------------------------------------------------------------------
# Tab 1: Lookup
# ---------------------------------------------------------------------------
lookup_frame = ttk.Frame(notebook, padding=10)
notebook.add(lookup_frame, text="Lookup")
lookup_frame.columnconfigure(1, weight=1)
lookup_frame.rowconfigure(0, weight=1)
lookup_frame.rowconfigure(1, weight=1)

lookup_form = ttk.Frame(lookup_frame)
lookup_form.grid(row=0, column=0, sticky="nw", padx=(0, 10))

spirit_label = tk.Label(lookup_form, text="Spirit Code:")
spirit_label.grid(row=0, column=0, sticky="e", padx=5, pady=5)
spirit_entry = tk.Entry(lookup_form, width=30)
spirit_entry.grid(row=0, column=1, padx=5, pady=5)

hotel_label = tk.Label(lookup_form, text="Hotel:")
hotel_label.grid(row=1, column=0, sticky="e", padx=5, pady=5)
hotel_var = tk.StringVar()
hotel_combo = ttk.Combobox(lookup_form, textvariable=hotel_var, values=hotel_names)
hotel_combo.grid(row=1, column=1, padx=5, pady=5)
hotel_combo.state(["!readonly"])


def on_hotel_keyrelease(event):
    val = hotel_var.get()
    hotel_combo["values"] = hotel_names if not val else [h for h in hotel_names if val.lower() in h.lower()]


hotel_combo.bind("<KeyRelease>", on_hotel_keyrelease)

search_button = tk.Button(lookup_form, text="Search", command=lambda: lookup(spirit_entry, hotel_var))
search_button.grid(row=2, column=0, columnspan=2, pady=10)

single_attach_frame = ttk.LabelFrame(lookup_form, text="Attachments (single email)", padding=8)
single_attach_frame.grid(row=3, column=0, columnspan=2, sticky="we", padx=2, pady=6)
single_attachments_enabled_var = tk.BooleanVar(value=False)
single_attachments_root_var = tk.StringVar(value="")
ttk.Checkbutton(single_attach_frame, text="Enable attachments for single email", variable=single_attachments_enabled_var).grid(row=0, column=0, sticky="w", padx=4, pady=2)
ttk.Label(single_attach_frame, text="Attachments root").grid(row=1, column=0, sticky="w", padx=4, pady=2)
single_attach_entry = ttk.Entry(single_attach_frame, textvariable=single_attachments_root_var, width=40)
single_attach_entry.grid(row=1, column=1, sticky="ew", padx=4, pady=2)

def browse_single_attach_root():
    sel = filedialog.askdirectory(title="Choose attachments root (single email)")
    if sel:
        single_attachments_root_var.set(sel)

ttk.Button(single_attach_frame, text="Browse", command=browse_single_attach_root).grid(row=1, column=2, sticky="e", padx=4, pady=2)
single_attach_frame.columnconfigure(1, weight=1)

detail_container = ttk.Frame(lookup_frame)
detail_container.grid(row=0, column=1, sticky="nsew")
init_detail_panel(detail_container)
clear_detail_panel()
init_single_compose_ui(lookup_frame)

# ---------------------------------------------------------------------------
# Tab 2: Multi-email
# ---------------------------------------------------------------------------
multi_frame = ttk.Frame(notebook, padding=10)
notebook.add(multi_frame, text="Multi-Email")

root.minsize(1150, 820)
forward_bar = ttk.Frame(multi_frame)
forward_bar.pack(fill="x", padx=5, pady=(0, 6))
forward_status_var = tk.StringVar(value="No source email captured.")
ttk.Button(forward_bar, text="Browse Outlook...", command=browse_outlook_email).pack(side="left", padx=4)
ttk.Button(forward_bar, text="Clear Forward", command=clear_forward_template).pack(side="left", padx=4)
ttk.Label(forward_bar, textvariable=forward_status_var, foreground="gray").pack(side="left", padx=8)

quick_frame = ttk.Frame(multi_frame)
quick_frame.pack(fill="x", padx=5, pady=(0, 6))
ttk.Label(quick_frame, text="Quick Spirit Codes (comma separated)").pack(side="left", padx=4)
quick_spirit_var = tk.StringVar()
quick_entry = ttk.Entry(quick_frame, textvariable=quick_spirit_var)
quick_entry.pack(side="left", padx=4, fill="x", expand=True)
ttk.Button(quick_frame, text="Apply Quick Filter", command=refresh_filtered_hotels).pack(side="left", padx=4)
filtered_count_var = tk.StringVar(value="Filtered: 0")
ttk.Label(quick_frame, textvariable=filtered_count_var, foreground="gray").pack(side="right", padx=4)

# Multi-email attachments controls (moved here for visibility)
attachments_frame = ttk.LabelFrame(multi_frame, text="Attachments (multi-email)", padding=6)
attachments_frame.pack(fill="x", padx=5, pady=(0, 6))
ttk.Checkbutton(attachments_frame, text="Enable attachments", variable=attachments_enabled_var).grid(row=0, column=0, sticky="w", padx=4, pady=2)
ttk.Label(attachments_frame, text="Attachments root").grid(row=1, column=0, sticky="w", padx=4, pady=2)
attach_entry = ttk.Entry(attachments_frame, textvariable=attachments_root_var)
attach_entry.grid(row=1, column=1, sticky="ew", padx=4, pady=2)
ttk.Button(attachments_frame, text="Browse", command=browse_attachments_root).grid(row=1, column=2, sticky="e", padx=4, pady=2)
attachments_frame.columnconfigure(1, weight=1)

filters_frame = ttk.LabelFrame(multi_frame, text="Filter Hotels", padding=10)
filters_frame.pack(fill="x", padx=5, pady=5)

hyatt_year_var = tk.StringVar(value="")
hyatt_year_mode_var = tk.StringVar(value="Any")

def make_multiselect(parent, label_text):
    wrap = ttk.Frame(parent)
    ttk.Label(wrap, text=label_text).pack(anchor="w")
    lb = tk.Listbox(wrap, selectmode="extended", height=6, exportselection=False)
    lb.pack(side="left", fill="both", expand=True)
    sb = ttk.Scrollbar(wrap, orient="vertical", command=lb.yview)
    sb.pack(side="right", fill="y")
    lb.config(yscrollcommand=sb.set)
    return wrap, lb

row_f = 0
brand_wrap, brand_listbox = make_multiselect(filters_frame, "Brand (multi-select)")
brand_wrap.grid(row=row_f, column=0, sticky="nsew", padx=4, pady=2)

band_wrap, brand_band_listbox = make_multiselect(filters_frame, "Brand Band")
band_wrap.grid(row=row_f, column=1, sticky="nsew", padx=4, pady=2)

region_wrap, region_listbox = make_multiselect(filters_frame, "Region")
region_wrap.grid(row=row_f, column=2, sticky="nsew", padx=4, pady=2)

relationship_wrap, relationship_listbox = make_multiselect(filters_frame, "Relationship")
relationship_wrap.grid(row=row_f, column=3, sticky="nsew", padx=4, pady=2)

country_wrap, country_listbox = make_multiselect(filters_frame, "Country/Area")
country_wrap.grid(row=row_f, column=4, sticky="nsew", padx=4, pady=2)

hyatt_wrap = ttk.Frame(filters_frame)
hyatt_wrap.grid(row=row_f, column=5, sticky="nw", padx=4, pady=2)
ttk.Label(hyatt_wrap, text="Hyatt Date (year)").pack(anchor="w")
hyatt_year_entry = ttk.Entry(hyatt_wrap, textvariable=hyatt_year_var, width=10)
hyatt_year_entry.pack(anchor="w", pady=(0, 2))
hyatt_mode_combo = ttk.Combobox(
    hyatt_wrap,
    textvariable=hyatt_year_mode_var,
    values=["Any", "Before", "Before/Equal", "Equal", "After/Equal", "After"],
    state="readonly",
    width=12,
)
hyatt_mode_combo.pack(anchor="w")

for col in range(5):
    filters_frame.columnconfigure(col, weight=1)

apply_filter_btn = ttk.Button(filters_frame, text="Apply Filter", command=refresh_filtered_hotels)
apply_filter_btn.grid(row=0, column=6, sticky="e", padx=8, pady=2)

reset_filter_btn = ttk.Button(filters_frame, text="Reset Filters", command=reset_filters)
reset_filter_btn.grid(row=0, column=7, sticky="e", padx=8, pady=2)

lists_pane = ttk.Panedwindow(multi_frame, orient="horizontal")
lists_pane.pack(fill="both", expand=True, padx=5, pady=5)

# Buttons between filters and panes
buttons_bar = ttk.Frame(multi_frame)
buttons_bar.pack(fill="x", padx=5, pady=(0, 5))
ttk.Button(buttons_bar, text="Select", command=add_selected_hotels).pack(side="left", padx=4)
ttk.Button(buttons_bar, text="Select All", command=add_all_filtered_hotels).pack(side="left", padx=4)
ttk.Button(buttons_bar, text="Remove", command=remove_selected_hotels).pack(side="left", padx=4)
ttk.Button(buttons_bar, text="Remove All", command=clear_selected_hotels).pack(side="left", padx=4)

filtered_frame = ttk.LabelFrame(lists_pane, text="Gefilterte Hotels", padding=5)
lists_pane.add(filtered_frame, weight=1)

filtered_tree = ttk.Treeview(
    filtered_frame,
    columns=("Spirit", "Hotel", "City", "Brand", "Brand Band", "Relationship", "Region", "Country"),
    show="headings",
    selectmode="extended",
)
filtered_xscroll = ttk.Scrollbar(filtered_frame, orient="horizontal", command=filtered_tree.xview)
filtered_tree.configure(xscrollcommand=filtered_xscroll.set)
for col, width in [
    ("Spirit", 70),
    ("Hotel", 220),
    ("City", 120),
    ("Brand", 120),
    ("Brand Band", 120),
    ("Relationship", 120),
    ("Region", 120),
    ("Country", 140),
]:
    filtered_tree.heading(col, text=col)
    filtered_tree.column(col, width=width, stretch=True)
filtered_tree.pack(fill="both", expand=True)
filtered_xscroll.pack(fill="x")
enable_treeview_sort(filtered_tree)

selected_frame = ttk.LabelFrame(lists_pane, text="Ausgewaehlte Hotels", padding=5)
lists_pane.add(selected_frame, weight=1)

selected_columns = ("Spirit", "Hotel", "Recipients", "AVP", "MD", "GM", "ENG", "DOF", "RES")
selected_tree = ttk.Treeview(selected_frame, columns=selected_columns, show="headings", selectmode="extended")
selected_xscroll = ttk.Scrollbar(selected_frame, orient="horizontal", command=selected_tree.xview)
selected_tree.configure(xscrollcommand=selected_xscroll.set)
for col, width in [
    ("Spirit", 60),      # 5 chars
    ("Hotel", 260),
    ("Recipients", 500),
    ("AVP", 30),
    ("MD", 30),
    ("GM", 30),
    ("ENG", 30),
    ("DOF", 30),
    ("RES", 30),
]:
    selected_tree.heading(col, text=col)
    if col in ("AVP", "MD", "GM", "ENG", "DOF", "RES", "Spirit"):
        selected_tree.column(col, width=width, anchor="center", stretch=False)
    else:
        selected_tree.column(col, width=width, stretch=True)
selected_tree.pack(fill="both", expand=True)
selected_xscroll.pack(fill="x")
enable_treeview_sort(selected_tree)

draft_btn = ttk.Button(multi_frame, text="Draft Emails in Outlook", command=draft_emails_for_selection)
draft_btn.pack(anchor="e", padx=8, pady=6)
collective_btn = ttk.Button(multi_frame, text="Draft ONE collective email", command=draft_collective_email)
collective_btn.pack(anchor="e", padx=8, pady=(0, 6))

# ---------------------------------------------------------------------------
# Tab 3: Excel Emails
# ---------------------------------------------------------------------------
excel_frame = ttk.Frame(notebook, padding=10)
notebook.add(excel_frame, text="Excel Emails")

excel_top = ttk.Frame(excel_frame)
excel_top.pack(fill="x", pady=(0, 10))
ttk.Label(excel_top, text="Excel-Datei fuer Emails laden:").pack(side="left", padx=4)
ttk.Button(excel_top, text="Datei laden", command=load_excel_email_file).pack(side="left", padx=4)
ttk.Label(excel_top, textvariable=excel_file_label_var, foreground="gray").pack(side="left", padx=8)

# Forward controls (Excel)
excel_forward_bar = ttk.Frame(excel_frame)
excel_forward_bar.pack(fill="x", padx=5, pady=(0, 6))
ttk.Label(excel_forward_bar, text="Forward-Email (optional):").pack(side="left", padx=4)
ttk.Button(excel_forward_bar, text="Browse Outlook...", command=browse_outlook_email).pack(side="left", padx=4)
ttk.Button(excel_forward_bar, text="Clear Forward", command=clear_forward_template).pack(side="left", padx=4)
ttk.Label(excel_forward_bar, textvariable=forward_status_var, foreground="gray").pack(side="left", padx=8)

# Attachments controls (Excel)
excel_attachments_frame = ttk.LabelFrame(excel_frame, text="Attachments (Excel emails)", padding=6)
excel_attachments_frame.pack(fill="x", padx=5, pady=(0, 6))
ttk.Checkbutton(excel_attachments_frame, text="Enable attachments", variable=attachments_enabled_var).grid(row=0, column=0, sticky="w", padx=4, pady=2)
ttk.Label(excel_attachments_frame, text="Attachments root").grid(row=1, column=0, sticky="w", padx=4, pady=2)
attach_entry_excel = ttk.Entry(excel_attachments_frame, textvariable=attachments_root_var)
attach_entry_excel.grid(row=1, column=1, sticky="ew", padx=4, pady=2)
ttk.Button(excel_attachments_frame, text="Browse", command=browse_attachments_root).grid(row=1, column=2, sticky="e", padx=4, pady=2)
excel_attachments_frame.columnconfigure(1, weight=1)

mapping_box = ttk.LabelFrame(excel_frame, text="Spaltenzuordnung", padding=8)
mapping_box.pack(fill="both", expand=True, pady=(0, 10))
ttk.Label(mapping_box, text="Waehlen Sie je Spalte: Spirit Code / Include in Body / Skip").pack(anchor="w", pady=(0, 6))
mapping_container = ttk.Frame(mapping_box)
mapping_container.pack(fill="both", expand=True)
headers_canvas = tk.Canvas(mapping_container, borderwidth=0, highlightthickness=0)
headers_scroll = ttk.Scrollbar(mapping_container, orient="vertical", command=headers_canvas.yview)
excel_headers_frame = ttk.Frame(headers_canvas)
headers_window = headers_canvas.create_window((0, 0), window=excel_headers_frame, anchor="nw")
def _on_headers_config(event):
    headers_canvas.configure(scrollregion=headers_canvas.bbox("all"))
excel_headers_frame.bind("<Configure>", _on_headers_config)
headers_canvas.configure(yscrollcommand=headers_scroll.set)
headers_canvas.pack(side="left", fill="both", expand=True)
headers_scroll.pack(side="right", fill="y")

filter_bar = ttk.Frame(excel_frame)
filter_bar.pack(fill="x", pady=4)
excel_filter_summary_var = tk.StringVar(value="")
ttk.Button(filter_bar, text="Filter anwenden", command=refresh_excel_filtered_tree).pack(side="left", padx=4)
ttk.Button(filter_bar, text="Filter loeschen", command=lambda: [lb.selection_clear(0, tk.END) for _, _, lb in excel_filter_controls] or refresh_excel_filtered_tree()).pack(side="left", padx=4)
ttk.Label(filter_bar, textvariable=excel_filter_summary_var, foreground="gray").pack(side="left", padx=6)
excel_filtered_count_var = tk.StringVar(value="Gefiltert: 0")
ttk.Label(filter_bar, textvariable=excel_filtered_count_var, foreground="gray").pack(side="left", padx=6)

actions = ttk.Frame(excel_frame)
actions.pack(fill="x", pady=4)
ttk.Label(actions, text="Versand-Modus:").pack(side="left", padx=10)
ttk.Radiobutton(actions, text="Einzel pro Hotel", variable=excel_mode_var, value="dedicated").pack(side="left")
ttk.Radiobutton(actions, text="Eine Sammelmail", variable=excel_mode_var, value="collective").pack(side="left", padx=6)
ttk.Button(actions, text="Emails erstellen", command=prompt_excel_compose).pack(side="right", padx=4)

excel_lists = ttk.Panedwindow(excel_frame, orient="horizontal")
excel_lists.pack(fill="both", expand=True, pady=6)

excel_filtered_frame = ttk.LabelFrame(excel_lists, text="Gefilterte Excel-Hotels", padding=5)
excel_lists.add(excel_filtered_frame, weight=1)
excel_filtered_tree = ttk.Treeview(
    excel_filtered_frame,
    columns=("Spirit", "Hotel", "Recipients", "AVP", "MD", "GM", "ENG", "DOF", "RES"),
    show="headings",
    selectmode="extended",
)
for col, width in [
    ("Spirit", 70),
    ("Hotel", 220),
    ("Recipients", 420),
    ("AVP", 40),
    ("MD", 40),
    ("GM", 40),
    ("ENG", 40),
    ("DOF", 40),
    ("RES", 40),
]:
    excel_filtered_tree.heading(col, text=col)
    if col in ("AVP", "MD", "GM", "ENG", "DOF", "RES", "Spirit"):
        excel_filtered_tree.column(col, width=width, anchor="center", stretch=False)
    else:
        excel_filtered_tree.column(col, width=width, stretch=True)
excel_filtered_tree.pack(fill="both", expand=True)
enable_treeview_sort(excel_filtered_tree)

excel_selected_frame = ttk.LabelFrame(excel_lists, text="Ausgewaehlte Excel-Hotels", padding=5)
excel_lists.add(excel_selected_frame, weight=1)
excel_selected_tree = ttk.Treeview(
    excel_selected_frame,
    columns=("Spirit", "Hotel", "Recipients", "AVP", "MD", "GM", "ENG", "DOF", "RES"),
    show="headings",
    selectmode="extended",
)
for col, width in [
    ("Spirit", 70),
    ("Hotel", 220),
    ("Recipients", 420),
    ("AVP", 40),
    ("MD", 40),
    ("GM", 40),
    ("ENG", 40),
    ("DOF", 40),
    ("RES", 40),
]:
    excel_selected_tree.heading(col, text=col)
    if col in ("AVP", "MD", "GM", "ENG", "DOF", "RES", "Spirit"):
        excel_selected_tree.column(col, width=width, anchor="center", stretch=False)
    else:
        excel_selected_tree.column(col, width=width, stretch=True)
excel_selected_tree.pack(fill="both", expand=True)
enable_treeview_sort(excel_selected_tree)

excel_btns = ttk.Frame(excel_frame)
excel_btns.pack(fill="x", pady=4)
ttk.Button(excel_btns, text="Auswahl hinzufuegen", command=excel_add_selected).pack(side="left", padx=4)
ttk.Button(excel_btns, text="Alle hinzufuegen", command=excel_add_all).pack(side="left", padx=4)
ttk.Button(excel_btns, text="Entfernen", command=excel_remove_selected).pack(side="left", padx=4)
ttk.Button(excel_btns, text="Alle entfernen", command=excel_clear_selected).pack(side="left", padx=4)

# ---------------------------------------------------------------------------
# Tab 4: Setup
# ---------------------------------------------------------------------------
setup_frame = ttk.Frame(notebook, padding=10)
notebook.add(setup_frame, text="Setup")

setup_top = ttk.LabelFrame(setup_frame, text="Data Columns", padding=10)
setup_top.pack(fill="x", padx=5, pady=5)

brand_col_combo = ttk.Combobox(setup_top, textvariable=brand_col_var, state="readonly")
region_col_combo = ttk.Combobox(setup_top, textvariable=region_col_var, state="readonly")
country_col_combo = ttk.Combobox(setup_top, textvariable=country_col_var, state="readonly")
country_fallback_combo = ttk.Combobox(setup_top, textvariable=country_fallback_col_var, state="readonly")
city_col_combo = ttk.Combobox(setup_top, textvariable=city_col_var, state="readonly")
brand_band_col_combo = ttk.Combobox(setup_top, textvariable=brand_band_col_var, state="readonly")
relationship_col_combo = ttk.Combobox(setup_top, textvariable=relationship_col_var, state="readonly")
hyatt_date_col_combo = ttk.Combobox(setup_top, textvariable=hyatt_date_col_var, state="readonly")

row_setup = 0
labels = [
    ("Brand column", brand_col_combo),
    ("Brand Band column", brand_band_col_combo),
    ("Region column", region_col_combo),
    ("Country column", country_col_combo),
    ("Country fallback (optional)", country_fallback_combo),
    ("City column", city_col_combo),
    ("Relationship column", relationship_col_combo),
    ("Hyatt Date column (for year filter)", hyatt_date_col_combo),
]
for idx, (text, combo) in enumerate(labels):
    ttk.Label(setup_top, text=text).grid(row=row_setup + idx, column=0, sticky="w", padx=5, pady=2)
    combo.grid(row=row_setup + idx, column=1, sticky="ew", padx=5, pady=2)
setup_top.columnconfigure(1, weight=1)

roles_setup = ttk.LabelFrame(setup_frame, text="Recipient Columns", padding=10)
roles_setup.pack(fill="x", padx=5, pady=5)

avp_col_combo = ttk.Combobox(roles_setup, textvariable=avp_col_var, state="readonly")
md_col_combo = ttk.Combobox(roles_setup, textvariable=md_col_var, state="readonly")
gm_col_combo = ttk.Combobox(roles_setup, textvariable=gm_col_var, state="readonly")
eng_col_combo = ttk.Combobox(roles_setup, textvariable=eng_col_var, state="readonly")
dof_col_combo = ttk.Combobox(roles_setup, textvariable=dof_col_var, state="readonly")
reg_eng_spec_combo = ttk.Combobox(roles_setup, textvariable=reg_eng_spec_col_var, state="readonly")

labels_roles = [
    ("AVP column", avp_col_combo),
    ("Managing Director column", md_col_combo),
    ("GM column", gm_col_combo),
    ("Engineering column", eng_col_combo),
    ("DOF column", dof_col_combo),
    ("Regional Eng Specialist column (optional)", reg_eng_spec_combo),
]
for idx, (text, combo) in enumerate(labels_roles):
    ttk.Label(roles_setup, text=text).grid(row=idx, column=0, sticky="w", padx=5, pady=2)
    combo.grid(row=idx, column=1, sticky="ew", padx=5, pady=2)
roles_setup.columnconfigure(1, weight=1)

role_delivery = ttk.LabelFrame(setup_frame, text="Role Delivery (To/CC/BCC)", padding=10)
role_delivery.pack(fill="x", padx=5, pady=5)
role_delivery.columnconfigure(1, weight=1)
add_role_selector(role_delivery, "AVP", "Skip")
add_role_selector(role_delivery, "MD", "Skip")
add_role_selector(role_delivery, "GM", "To")
add_role_selector(role_delivery, "Engineering", "CC")
add_role_selector(role_delivery, "DOF", "CC")
add_role_selector(role_delivery, "RegionalEngineeringSpecialist", "CC")

# browse function defined earlier; controls now on Multi-email tab only

visible_cols_frame = ttk.LabelFrame(setup_frame, text='Columns shown in "Gefilterte Hotels"', padding=10)
visible_cols_frame.pack(fill="both", padx=5, pady=5)
filter_cols_listbox = tk.Listbox(visible_cols_frame, selectmode="extended", height=8, exportselection=False)
filter_cols_listbox.pack(fill="both", expand=True)

apply_columns_btn = ttk.Button(setup_frame, text="Apply column mapping", command=apply_column_settings)
apply_columns_btn.pack(anchor="e", padx=5, pady=10)

# Status bar
status_label = ttk.Label(root, textvariable=status_var, relief=tk.SUNKEN, anchor="w")
status_label.pack(fill="x", side="bottom")

# Initial startup flow (no auto data load). Show splash first, then close it before prompting for config.
ensure_initial_data()
# Note: splash is auto-closed after 2 minutes or via the "Understood..." button.
# Config prompt appears once the splash is closed.

# Populate setup dropdown values after any config/data load
refresh_setup_tab_options()

# Populate filter dropdown values after data load
update_filter_options()

# Initial filtered view
refresh_filtered_hotels()

# Auto-close splash after 2 minutes
root.after(120000, close_splash)

# Warm Outlook in the background so the first email opens faster
root.after(200, warm_outlook_app)

root.mainloop()
