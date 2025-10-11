"""Hilfsprogramm zur Konvertierung einer Excel-Tabelle in die Spirit-Fixture."""

from __future__ import annotations

import argparse
import json
import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Sequence, Tuple

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ModuleNotFoundError as exc:  # pragma: no cover - wird in Tests übersprungen
    raise ModuleNotFoundError(
        "Für den Excel-Import wird 'openpyxl' benötigt. Installiere das Paket z. B. mit `pip install openpyxl`."
    ) from exc


_DEFAULT_CONFIG_PATH = Path(__file__).resolve().parent.parent / "data" / "excel_helper_config.json"

_EMAIL_PATTERN = re.compile(r"email", re.IGNORECASE)
_MAIL_ALIAS_PATTERN = re.compile(r"mail", re.IGNORECASE)


def normalize_key(value: str) -> str:
    """Normalisiere Spaltenüberschriften für Vergleiche."""

    return re.sub(r"[^0-9a-z]", "", value.lower())


def is_empty(value: Any) -> bool:
    """Prüfe, ob eine Zelle leer oder nur Whitespace enthält."""

    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    if isinstance(value, float):
        return not math.isfinite(value)
    return False


def format_cell(value: Any) -> str | None:
    """Bereite Zellwerte für die JSON-Ausgabe auf."""

    if is_empty(value):
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if not math.isfinite(value):
            return None
        if value.is_integer():
            return str(int(value))
        return format(value, "g")
    return str(value).strip() or None


def load_rows(sheet: Worksheet) -> Tuple[List[str], List[List[Any]]]:
    """Lese die Header und Datenzeilen aus einem Worksheet."""

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    data_rows = [list(row) for row in rows[1:]]
    return headers, data_rows


def _looks_like_email(label: str) -> bool:
    normalized = normalize_key(label)
    return bool(_EMAIL_PATTERN.search(label) or _MAIL_ALIAS_PATTERN.search(label) or "mail" in normalized)


def detect_email_columns(headers: Sequence[str]) -> List[str]:
    """Ermittle plausible E-Mail-Spalten aus den Überschriften."""

    detected: List[str] = []
    for header in headers:
        if not header:
            continue
        if _looks_like_email(header):
            detected.append(header)
    return detected


_FIELD_ALIASES: Dict[str, List[str]] = {
    "spiritCode": ["spiritcode"],
    "displayField": [
        "hotelname",
        "hotel",
        "propertyname",
        "property",
        "projektname",
        "projectname",
        "assetname",
        "standortname",
        "locationname",
        "site",
    ],
    "region": ["region"],
    "status": ["status", "projektstatus", "pipelinestatus"],
    "city": ["city", "ort", "stadt", "locationcity"],
    "country": ["country", "land", "locationcountry"],
    "address": ["address", "adresse", "street", "strasse", "locationaddress"],
}


def derive_field_mapping(columns: Sequence[str]) -> Dict[str, str]:
    """Leite eine Zuordnung bekannter Felder aus den Spaltennamen ab."""

    mapping: Dict[str, str] = {}
    used: set[str] = set()
    normalized = {column: normalize_key(column) for column in columns}

    for field, aliases in _FIELD_ALIASES.items():
        for column, key in normalized.items():
            if column in used:
                continue
            if key in aliases:
                mapping[field] = column
                used.add(column)
                break

    # Sicherstellen, dass eine Anzeige-Spalte vorhanden ist
    if "displayField" not in mapping:
        for column in columns:
            if column in used:
                continue
            normalized_key = normalize_key(column)
            if _looks_like_email(column):
                continue
            if normalized_key.endswith("phone") or normalized_key.endswith("telefon"):
                continue
            mapping["displayField"] = column
            used.add(column)
            break

    return mapping


@dataclass(frozen=True)
class HelperConfig:
    selected_columns: List[str]
    email_columns: List[str]


def load_helper_config(excel_path: Path, config_path: Path | None = None) -> HelperConfig | None:
    """Lade die gespeicherte Auswahl aus der Excel-Helper-Konfiguration."""

    path = config_path or _DEFAULT_CONFIG_PATH
    if not path.exists():
        return None
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return None
    if not isinstance(raw, dict):
        return None
    files = raw.get("files")
    if not isinstance(files, dict):
        return None

    resolved = str(excel_path.resolve())
    entry = files.get(resolved)
    if entry is None:
        # Fallback: versuche anhand des Dateinamens einen Treffer zu finden
        matches = [value for key, value in files.items() if Path(key).name == excel_path.name]
        if len(matches) == 1:
            entry = matches[0]
    if not isinstance(entry, dict):
        return None

    selected = entry.get("selectedColumns") or []
    email = entry.get("emailColumns") or []

    selected_columns = [str(item) for item in selected if isinstance(item, str)]
    email_columns = [str(item) for item in email if isinstance(item, str)]

    if not selected_columns:
        return None
    return HelperConfig(selected_columns=selected_columns, email_columns=email_columns)


def _resolve_selected_columns(
    headers: Sequence[str], helper: HelperConfig | None
) -> Tuple[List[str], List[str], List[str]]:
    """Ermittle die zu verwendenden Spalten und zusätzliche Warnungen."""

    warnings: List[str] = []
    if helper:
        missing = [column for column in helper.selected_columns if column not in headers]
        if missing:
            missing_list = ", ".join(missing)
            raise ValueError(
                f"Die gespeicherte Konfiguration enthält unbekannte Spalten: {missing_list}. Bitte prüfen Sie die Excel-Datei."
            )
        selected = list(helper.selected_columns)
    else:
        selected = [header for header in headers if header]

    email_columns: List[str]
    if helper and helper.email_columns:
        email_columns = [column for column in helper.email_columns if column in selected]
        ignored = [column for column in helper.email_columns if column not in selected]
        if ignored:
            warnings.append(
                "E-Mail-Spalten aus der Konfiguration wurden ignoriert, da sie nicht ausgewählt wurden: "
                + ", ".join(ignored)
            )
    else:
        email_columns = detect_email_columns(selected)

    return selected, email_columns, warnings


def _build_field_index(headers: Sequence[str]) -> Dict[str, int]:
    index: Dict[str, int] = {}
    for idx, header in enumerate(headers):
        if header and header not in index:
            index[header] = idx
    return index


def convert_excel_to_fixture(
    excel_path: Path,
    *,
    sheet_name: str | None = None,
    config_path: Path | None = None,
) -> Tuple[Dict[str, Any], List[str]]:
    """Konvertiere eine Excel-Datei in eine strukturierte Fixture."""

    workbook = load_workbook(excel_path, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"Arbeitsblatt '{sheet_name}' existiert nicht. Verfügbare Blätter: {', '.join(workbook.sheetnames)}"
                )
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.active

        headers, rows = load_rows(worksheet)
    finally:
        workbook.close()

    helper = load_helper_config(excel_path, config_path=config_path)
    selected_columns, email_columns, warnings = _resolve_selected_columns(headers, helper)
    if not selected_columns:
        raise ValueError("Es wurden keine Spalten zur Konvertierung gefunden.")

    field_mapping = derive_field_mapping(selected_columns)
    spirit_column = field_mapping.get("spiritCode")
    if not spirit_column:
        raise ValueError(
            "Erforderliche Spalte 'Spirit Code' konnte nicht automatisch ermittelt werden."
            " Bitte stellen Sie sicher, dass eine entsprechende Spalte vorhanden ist."
        )

    column_index = _build_field_index(headers)

    records: List[Dict[str, Any]] = []
    for offset, row in enumerate(rows, start=2):
        if all(is_empty(value) for value in row):
            continue

        fields: Dict[str, str | None] = {}
        for column in selected_columns:
            idx = column_index.get(column)
            cell_value = row[idx] if idx is not None and idx < len(row) else None
            fields[column] = format_cell(cell_value)

        spirit_value = fields.get(spirit_column)
        if not spirit_value:
            raise ValueError(f"Zeile {offset}: Die erforderliche Spalte '{spirit_column}' enthält keinen Wert.")

        display_column = field_mapping.get("displayField")
        display_value = fields.get(display_column) if display_column else None
        if not display_value:
            display_value = spirit_value

        emails = {
            column: value
            for column in email_columns
            if (value := fields.get(column))
        }

        record_entry: Dict[str, Any] = {
            "spiritCode": spirit_value,
            "displayValue": display_value,
            "fields": fields,
        }
        if emails:
            record_entry["emails"] = emails
        records.append(record_entry)

    fixture = {
        "config": {
            "selectedColumns": selected_columns,
            "emailColumns": email_columns,
            "fieldMapping": field_mapping,
        },
        "records": records,
    }
    return fixture, warnings


def write_fixture(fixture: Dict[str, Any], output_path: Path, *, indent: int = 2) -> None:
    """Schreibe die Fixture-Daten in eine JSON-Datei."""

    output_path.write_text(
        json.dumps(fixture, indent=indent, ensure_ascii=False),
        encoding="utf-8",
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Konvertiert eine Excel-Datei in die Spirit Lookup JSON-Fixture",
    )
    parser.add_argument("excel_path", type=Path, help="Pfad zur Excel-Datei (.xlsx)")
    parser.add_argument(
        "output_path",
        type=Path,
        nargs="?",
        help="Zieldatei für die JSON-Fixture (Standard: gleicher Pfad mit .json)",
    )
    parser.add_argument(
        "--sheet",
        dest="sheet_name",
        help="Name des Arbeitsblatts, das konvertiert werden soll (Standard: erstes Blatt)",
    )
    parser.add_argument(
        "--config",
        dest="config_path",
        type=Path,
        help="Pfad zur Excel-Helper-Konfiguration (Standard: data/excel_helper_config.json)",
    )
    parser.add_argument(
        "--indent",
        type=int,
        default=2,
        help="Einrückung im JSON-Output (Standard: 2)",
    )
    parser.add_argument(
        "--fail-on-warning",
        action="store_true",
        help="Beende mit Fehler, wenn Spalten nicht zugeordnet werden konnten",
    )
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    excel_path: Path = args.excel_path
    if not excel_path.exists():
        parser.error(f"Die Datei '{excel_path}' wurde nicht gefunden.")

    output_path: Path = args.output_path or excel_path.with_suffix(".json")

    try:
        fixture, warnings = convert_excel_to_fixture(
            excel_path,
            sheet_name=args.sheet_name,
            config_path=args.config_path,
        )
    except ValueError as exc:
        parser.error(str(exc))
        return 2

    write_fixture(fixture, output_path, indent=args.indent)

    message_lines = [
        f"✅ {len(fixture.get('records', []))} Datensätze nach '{output_path}' geschrieben.",
    ]
    if warnings:
        warning_block = "\n".join(f"⚠️  {warning}" for warning in warnings)
        if args.fail_on_warning:
            parser.error(f"Es wurden Warnungen erzeugt:\n{warning_block}")
        message_lines.append("Warnungen:")
        message_lines.append(warning_block)

    print("\n".join(message_lines))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

